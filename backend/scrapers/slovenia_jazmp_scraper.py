"""
Slovenia JAZMP Marketed Medicinal Products Scraper
----------------------------------------------------
Source:  JAZMP -- Javna agencija Republike Slovenije za zdravila in
         medicinske pripomocke (Public Agency of the Republic of Slovenia
         for Medicinal Products and Medical Devices)
URL:     https://www.jazmp.si/en/human-medicines/data-on-medicinal-products/marketed-medicinal-products/

JAZMP publishes an Excel workbook listing every notice a marketing
authorisation holder (MAH) has filed about a medicine's presence on the
Slovenian market: supply disruptions, temporary/permanent discontinuations,
and (re-)arrivals on the wholesale market. The English page is a thin
wrapper -- the underlying file and its contents are entirely in Slovenian.

The page links to both a ".pdf" (232 pages, human-readable) and an ".xls"
export. In practice the linked ".xls" URL 404s (the actual file behind it
is served as ".xlsx", not ".xls" -- a stale extension on JAZMP's own page)
so fetch() discovers the real "Prisotnost" attachment href(s) from the page
HTML each run and tries ".xlsx" as a fallback before giving up, rather than
hardcoding a brittle absolute file URL.

The workbook has two sheets, both with the same 10 columns:
    "Zadnja obvestila <date>" (Latest notices) -- ONE row per drug
        registration number (DS zdravila), holding only its most recent
        notice. This is the "current status" view and is what we use.
    "Vsa obvestila <date>" (All notices) -- full historical log (every
        notice ever filed), ~3x more rows. Not used here; a future pass
        could mine it for onset/resolution pairs per case number (Zadeva).

Columns (Slovenian header -> meaning):
    DS zdravila               Drug registration number (unique per pack)
    Zadeva                    Case/matter number, e.g. "2026-0947"
    Zdravilo                  Full product string: brand + strength + form
                               + pack size (Slovenian), e.g.
                               "Ursofalk 250 mg trde kapsule, skatla s 50 ..."
    Porocevalec                Reporting company (MAH) name
    Status porocevalca          Reporter's legal status (MAH / temp-MAH / etc)
    Datum prejema obvestila    Date JAZMP received the notice
    Vrsta obvestila            Notice type (see _classify_notice below)
    Datum zacetka              Start date of the reported event
    Napovedani datum konca     Announced/expected end date (nullable)
    Opombe JAZMP               JAZMP's free-text notes (nullable)

Notice-type taxonomy (Vrsta obvestila), classified by Slovenian keyword:
    "motnj[ai] v preskrbi"                      -> DISRUPTION (active shortage)
        preskrba = supply; motnja = disruption
    "zacasnega/zacasnem prenehanj[ae]"          -> TEMPORARY_STOP
        zacasno prenehanje (opravljanja prometa) = temporary cessation
        of marketing
    "stalnega/stalnem prenehanj[ae]"            -> PERMANENT_STOP
        stalno prenehanje (opravljanja prometa) = permanent cessation
        of marketing (= discontinuation)
    "prihod" / "zacetka prometa" / "pricetka"   -> ARRIVAL (incl. "ponovni
        prihod" = re-arrival, i.e. the medicine has returned to market
        after a previously-notified disruption/stop) -> resolved

Status mapping used here:
    DISRUPTION       -> 'active' (or 'anticipated' if the reported start
                         date for the disruption is still in the future)
    TEMPORARY_STOP    -> 'active' (temporary market withdrawal, not yet
                         resolved) / 'anticipated' if start date is future
    PERMANENT_STOP    -> 'resolved' (treated as a closed/discontinuation
                         event -- there is nothing more JAZMP will report
                         for this pack) with reason_category=discontinuation
    ARRIVAL/RETURN    -> 'resolved' (medicine is back on the market)

To avoid dumping ~9,000 historical "latest notice per drug since 2008"
rows (most of which are ancient, closed events with no ongoing signal),
normalize() only emits:
    - ALL currently-open disruptions/temporary stops (status active/
      anticipated) regardless of age, since these represent live shortage
      state per JAZMP's own record, and
    - Resolved/discontinuation notices from roughly the last 24 months,
      to keep a recent history trail without importing the full 18-year
      backlog on every run (MD5 dedup makes re-running safe either way,
      but there is no analytical value in ingesting a 2008 discontinuation
      as new "recent" scraper output).

Data source UUID:  10000000-0000-0000-0000-000000000104
Country:           Slovenia
Country code:      SI
Cron:              Weekly (scrape_frequency_hours=168 per migration 064)
"""

from __future__ import annotations

import io
import re
from datetime import date, datetime, timedelta
from typing import Any

import httpx

from backend.scrapers.base_scraper import BaseScraper, ScraperError
from backend.utils.reason_mapper import map_reason_category

# How far back to keep resolved/discontinuation notices (see module docstring).
_RESOLVED_LOOKBACK_DAYS = 730

# Sheet name prefixes (JAZMP appends the export date, e.g. "Zadnja obvestila
# 29-jun-2026" -- match on prefix so the date suffix doesn't break us).
_LATEST_SHEET_PREFIX = "Zadnja obvestila"
_ALL_SHEET_PREFIX = "Vsa obvestila"

# Dosage-form / packaging keywords (Slovenian) used to cut a generic name out
# of the full "Zdravilo" product string, which is brand + strength + form +
# pack, e.g. "Ursofalk 250 mg trde kapsule, skatla s 50 kapsulami ...".
_FORM_STOPWORDS = {
    "mg", "mg/ml", "g", "g/ml", "ml", "mcg", "microgramov", "ie", "%",
    "tablete", "tableta", "tablet", "kapsule", "kapsula", "kapsul",
    "praske", "prasek", "praska", "granulat", "zrnca", "raztopina",
    "suspenzija", "emulzija", "gel", "krema", "mazilo", "obliz", "obliz,",
    "sirup", "kapljice", "pastile", "sprej", "inhalacije", "vcepilo",
    "koncentrat", "liofilizat", "svecke", "vaginalete", "filmsko",
    "obloz ene", "obloz eno", "trde", "trdi", "podaljsanim", "sproscanjem",
    "prasek/koncentrat", "raztopino", "injiciranje", "infundiranje",
    "skatla", "vsebnik", "plastenko", "steklenico",
}

# Slovenian shortage/discontinuation reason vocabulary, layered on top of
# the centralised map_reason_category() free-text matcher.
_SL_REASON_MAP: dict[str, str] = {
    "motnja v preskrbi": "supply_chain",
    "motnji v preskrbi": "supply_chain",
    "prenehanje opravljanja prometa": "discontinuation",
    "stalno prenehanje": "discontinuation",
    "stalnega prenehanja": "discontinuation",
    "zacasno prenehanje": "supply_chain",
    "zacasnega prenehanja": "supply_chain",
    "proizvodne tezave": "manufacturing_issue",
    "tezave pri proizvodnji": "manufacturing_issue",
    "pomanjkanje surovin": "raw_material",
    "povecano povprasevanje": "demand_surge",
    "regulativni razlog": "regulatory_action",
}


class SloveniaJAZMPScraper(BaseScraper):
    SOURCE_ID: str    = "10000000-0000-0000-0000-000000000104"
    SOURCE_NAME: str  = "JAZMP — Marketed Medicinal Products (Slovenia)"
    BASE_URL: str     = (
        "https://www.jazmp.si/en/human-medicines/data-on-medicinal-products/"
        "marketed-medicinal-products/"
    )
    COUNTRY: str      = "Slovenia"
    COUNTRY_CODE: str = "SI"

    RATE_LIMIT_DELAY: float = 3.0   # Polite to a small national agency server
    REQUEST_TIMEOUT: float  = 90.0  # The workbook is a multi-MB download
    SCRAPER_VERSION: str    = "1.0.0"

    # -------------------------------------------------------------------------
    # fetch()
    # -------------------------------------------------------------------------

    def fetch(self) -> list[dict]:
        """
        Fetch JAZMP's marketed-medicinal-products workbook.

        Strategy:
        1. GET the (English-wrapper) landing page.
        2. Parse HTML for the "Prisotnost" attachment link(s) -- JAZMP
           links a ".pdf" and an ".xls" of the same underlying report.
        3. Download the workbook. The page's ".xls" href routinely 404s;
           the real file lives at the same path with an ".xlsx" extension,
           so that is tried as a fallback before giving up on this run.
        4. Parse the "Zadnja obvestila ..." (latest-notice-per-drug) sheet
           with openpyxl into a list of raw row dicts.
        """
        from bs4 import BeautifulSoup

        self.log.info("Scrape started", extra={
            "source": self.SOURCE_NAME,
            "url": self.BASE_URL,
        })

        resp = self._get(self.BASE_URL)
        soup = BeautifulSoup(resp.text, "html.parser")

        workbook_url = self._find_workbook_url(soup)
        if not workbook_url:
            # Raise, don't return []: a silent 0-record "success" hashes as a
            # duplicate on every later run, refreshing last_verified_at and
            # re-activating stale SI events — poisoning the freshness surface.
            raise ScraperError(
                f"No Excel/xls attachment link found on JAZMP page ({self.BASE_URL})"
            )

        content = self._download_workbook(workbook_url)
        if content is None:
            raise ScraperError(
                f"Could not download JAZMP workbook (xls/xlsx both failed): {workbook_url}"
            )

        records = self._parse_workbook(content)
        self.log.info("JAZMP fetch complete", extra={"records": len(records)})
        return records

    def _find_workbook_url(self, soup) -> str | None:
        """Find the 'Prisotnost' xls/xlsx attachment link on the landing page."""
        candidates: list[str] = []
        for link in soup.find_all("a", href=True):
            href = link["href"]
            href_lower = href.lower()
            if href_lower.endswith(".xls") or href_lower.endswith(".xlsx"):
                if not href.startswith("http"):
                    href = f"https://www.jazmp.si{href}" if href.startswith("/") \
                        else f"https://www.jazmp.si/{href}"
                candidates.append(href)

        if not candidates:
            return None

        # Prefer a link whose filename hints at the disruption/discontinuation
        # list ("prenehanja"/"motnje") over unrelated xls attachments
        # elsewhere on the page (e.g. the EMA critical-medicines list).
        for href in candidates:
            low = href.lower()
            if "prenehanja" in low or "motnje" in low or "prisotnost" in low:
                return href
        return candidates[0]

    def _download_workbook(self, url: str) -> bytes | None:
        """Download the workbook, falling back .xls -> .xlsx on 404."""
        urls_to_try = [url]
        if url.lower().endswith(".xls"):
            urls_to_try.append(url[:-4] + ".xlsx")

        for candidate in urls_to_try:
            try:
                resp = self._get(candidate)
                return resp.content
            except httpx.HTTPStatusError as exc:
                self.log.info(
                    "Workbook URL failed, trying next candidate",
                    extra={"url": candidate, "status": exc.response.status_code},
                )
                continue
            except Exception as exc:
                self.log.warning(
                    "Workbook download error",
                    extra={"url": candidate, "error": str(exc)},
                )
                continue
        return None

    def _parse_workbook(self, content: bytes) -> list[dict]:
        """Parse the 'Zadnja obvestila' (latest-per-drug) sheet into raw dicts."""
        try:
            import openpyxl
        except ImportError as exc:
            # Missing dep must fail loudly — a silent [] would poison freshness.
            raise ScraperError(
                "openpyxl not installed -- required for JAZMP xlsx parsing. "
                "Install with: pip install openpyxl"
            ) from exc

        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        except Exception as exc:
            raise ScraperError(f"Failed to open JAZMP workbook: {exc}") from exc

        sheet_name = next(
            (name for name in wb.sheetnames if name.startswith(_LATEST_SHEET_PREFIX)),
            None,
        )
        if sheet_name is None:
            # Fall back to the first sheet if JAZMP renames the tab.
            self.log.warning(
                "Could not find 'Zadnja obvestila' sheet by name; using first sheet",
                extra={"sheets": wb.sheetnames},
            )
            sheet_name = wb.sheetnames[0] if wb.sheetnames else None
        if sheet_name is None:
            return []

        ws = wb[sheet_name]
        rows = ws.iter_rows(min_row=2, values_only=True)

        records: list[dict] = []
        for row in rows:
            if not row or all(v is None for v in row):
                continue
            # Columns: DS zdravila, Zadeva, Zdravilo, Porocevalec,
            #          Status porocevalca, Datum prejema obvestila,
            #          Vrsta obvestila, Datum zacetka,
            #          Napovedani datum konca, Opombe JAZMP
            padded = list(row) + [None] * (10 - len(row))
            (
                reg_no, case_no, product, reporter, reporter_status,
                notice_date, notice_type, start_date, end_date, notes,
            ) = padded[:10]

            if not product or not notice_type:
                continue

            records.append({
                "registration_no": str(reg_no).strip() if reg_no else "",
                "case_no":         str(case_no).strip() if case_no else "",
                "product":         str(product).strip(),
                "reporter":        str(reporter).strip() if reporter else "",
                "reporter_status": str(reporter_status).strip() if reporter_status else "",
                "notice_date":     notice_date,
                "notice_type":     str(notice_type).strip(),
                "start_date":      start_date,
                "end_date":        end_date,
                "notes":           str(notes).strip() if notes else "",
                "sheet":           sheet_name,
            })

        self.log.info(
            "Parsed JAZMP workbook",
            extra={"sheet": sheet_name, "rows": len(records)},
        )
        return records

    # -------------------------------------------------------------------------
    # normalize()
    # -------------------------------------------------------------------------

    def normalize(self, raw: list[dict]) -> list[dict]:
        """Normalize JAZMP records into standard shortage event dicts."""
        self.log.info(
            "Normalising JAZMP records",
            extra={"source": self.SOURCE_NAME, "raw_count": len(raw)},
        )

        normalised: list[dict] = []
        skipped = 0
        today = date.today()
        lookback_cutoff = today - timedelta(days=_RESOLVED_LOOKBACK_DAYS)

        for rec in raw:
            try:
                result = self._normalise_record(rec, today, lookback_cutoff)
                if result is None:
                    skipped += 1
                    continue
                normalised.append(result)
            except Exception as exc:
                skipped += 1
                self.log.warning(
                    "Failed to normalise JAZMP record",
                    extra={"error": str(exc), "rec": str(rec)[:200]},
                )

        self.log.info(
            "Normalisation done",
            extra={"total": len(raw), "normalised": len(normalised), "skipped": skipped},
        )
        return normalised

    def _normalise_record(
        self, rec: dict, today: date, lookback_cutoff: date,
    ) -> dict | None:
        """Convert a single JAZMP row into a normalised shortage event dict."""
        product = rec.get("product", "").strip()
        if not product:
            return None

        notice_type = rec.get("notice_type", "")
        notice_class = self._classify_notice(notice_type, rec.get("notes", ""))

        event_start = self._parse_date(rec.get("start_date"))
        notice_date = self._parse_date(rec.get("notice_date"))
        event_end = self._parse_date(rec.get("end_date"))

        status = self._determine_status(notice_class, event_start, today)

        # Filter out ancient, already-closed events -- see module docstring.
        # Keep everything currently active/anticipated regardless of age
        # (JAZMP's own "latest notice" record says it's still open).
        if status in ("resolved",):
            anchor = event_start or notice_date
            if anchor:
                try:
                    anchor_date = datetime.fromisoformat(anchor).date()
                except ValueError:
                    anchor_date = today
                if anchor_date < lookback_cutoff:
                    return None

        generic_name = self._extract_generic_name(product)
        if not generic_name:
            return None

        reason_category, reason_text = self._map_reason(notice_class, notice_type)

        severity = "medium"
        if notice_class == "PERMANENT_STOP":
            severity = "high"
        elif notice_class == "DISRUPTION":
            severity = "medium"

        notes_parts: list[str] = []
        reporter = rec.get("reporter", "")
        if reporter:
            notes_parts.append(f"Reporter: {reporter}")
        notes_parts.append(f"Notice type (SL): {notice_type}")
        case_no = rec.get("case_no", "")
        if case_no:
            notes_parts.append(f"Case: {case_no}")
        jazmp_notes = rec.get("notes", "")
        if jazmp_notes:
            notes_parts.append(f"JAZMP notes: {jazmp_notes}")
        notes = "; ".join(notes_parts) or None

        anticipated_start_date = None
        if status == "anticipated":
            anticipated_start_date = event_start

        return {
            "generic_name":             generic_name.title(),
            "brand_names":              [product[:200]],
            "status":                   status,
            "severity":                 severity,
            "reason":                   reason_text,
            "reason_category":          reason_category,
            "start_date":               event_start or notice_date or today.isoformat(),
            "end_date":                 event_end if status == "resolved" else None,
            "estimated_resolution_date": event_end if status != "resolved" else None,
            "anticipated_start_date":   anticipated_start_date,
            "source_url":               self.BASE_URL,
            "notes":                    notes,
            "source_confidence_score":  85,
            "raw_record":               rec,
        }

    # -------------------------------------------------------------------------
    # Private helpers
    # -------------------------------------------------------------------------

    @staticmethod
    def _classify_notice(notice_type: str, notes: str = "") -> str:
        """
        Classify a Slovenian 'Vrsta obvestila' (notice type) string into one
        of the four buckets documented in the module docstring.

        Matching is done on a diacritic-folded, lowercased copy of the
        string so "začasnega"/"zacasnega" etc. match regardless of the
        exact accent characters JAZMP used in a given export.

        IMPORTANT #1: "prihod" (arrival / re-arrival on the market) must be
        checked FIRST for most cases. Re-arrival notice types read like
        "Ponovni prihod zdravila na trg na debelo po predhodno priglaseni
        motnji v preskrbi" ("Repeat arrival of the medicine on the wholesale
        market following a previously notified supply disruption") -- they
        contain "motnji"/"prenehanju" as a trailing reference to what
        preceded the arrival, not a description of the current notice.
        Checking "motnj"/"prenehanj" first would misclassify every
        re-arrival as still-open.

        IMPORTANT #2: one legacy notice type, "datum pricetka aktualnega
        zacasnega ali stalnega prenehanja opravljanja prometa z zdravilom"
        ("date of onset of the CURRENT temporary or permanent cessation of
        marketing"), contains "pricetka" (onset) but describes an ONGOING
        stop, not an arrival -- the opposite of "datum zacetka prometa z
        zdravilom" ("date marketing STARTED", i.e. a genuine arrival). Both
        strings are near-identical after folding, so this one case is
        disambiguated using JAZMP's free-text 'Opombe JAZMP' notes column,
        which spells out "zacasno prenehanje" / "stalno prenehanje" when
        this notice type is used.
        """
        low = _fold_diacritics(notice_type.lower())
        notes_low = _fold_diacritics((notes or "").lower())

        if "pricetka aktualnega" in low:
            # Legacy onset-of-current-stop notice -- NOT an arrival.
            # Disambiguate permanent vs temporary via the notes column;
            # default to temporary (the more common, less severe reading)
            # if the notes don't say either way.
            if "stalno prenehanje" in notes_low:
                return "PERMANENT_STOP"
            return "TEMPORARY_STOP"

        if "prihod" in low or "zacetka prometa" in low:
            return "ARRIVAL"
        if "motnj" in low:
            return "DISRUPTION"
        if "stalnega prenehanj" in low or "stalnem prenehanj" in low or "stalno prenehanje" in low:
            return "PERMANENT_STOP"
        if "zacasnega prenehanj" in low or "zacasnem prenehanj" in low or "zacasno prenehanje" in low:
            return "TEMPORARY_STOP"
        return "OTHER"

    @staticmethod
    def _determine_status(notice_class: str, event_start: str | None, today: date) -> str:
        """Map a notice classification (+ start date) to a shortage status."""
        is_future = False
        if event_start:
            try:
                is_future = datetime.fromisoformat(event_start).date() > today
            except ValueError:
                is_future = False

        if notice_class in ("DISRUPTION", "TEMPORARY_STOP"):
            return "anticipated" if is_future else "active"
        # PERMANENT_STOP and ARRIVAL/RETURN both represent a closed chapter
        # from JAZMP's perspective -- the medicine either left the market
        # for good, or came back. Either way there's nothing "ongoing".
        return "resolved"

    def _map_reason(self, notice_class: str, notice_type: str) -> tuple[str, str | None]:
        """
        Return (reason_category, reason_text) for a notice.

        notice_class is checked FIRST, not the raw notice_type text --
        arrival/re-arrival notices routinely reference the disruption or
        stop that preceded them ("... po predhodno priglaseni motnji v
        preskrbi" = "... following a previously notified supply
        disruption"), so a plain keyword scan over notice_type would
        mislabel a resolved re-arrival as an open supply_chain issue.
        """
        if notice_class == "PERMANENT_STOP":
            return "discontinuation", notice_type
        if notice_class in ("DISRUPTION", "TEMPORARY_STOP"):
            low = _fold_diacritics(notice_type.lower())
            for sl_phrase, category in _SL_REASON_MAP.items():
                if _fold_diacritics(sl_phrase) in low:
                    return category, notice_type
            return "supply_chain", notice_type
        if notice_class == "ARRIVAL":
            return "other", notice_type

        # OTHER / unclassified -- fall back to the centralised free-text
        # matcher (handles any additional English/French/etc. tokens that
        # slip through, e.g. JAZMP notes written partially in English).
        mapped = map_reason_category(notice_type)
        return mapped, notice_type or None

    @staticmethod
    def _extract_generic_name(product: str) -> str:
        """
        Heuristically extract a brand/molecule name from the JAZMP
        'Zdravilo' product string, which is brand + strength + form + pack,
        e.g.:
            "Ursofalk 250 mg trde kapsule, skatla s 50 kapsulami ..." -> "Ursofalk"
            "Rokuronijev bromid hameln 10 mg/ml raztopina ..."       -> "Rokuronijev Bromid Hameln"
            "SOMATULINE Autogel 60 mg raztopina za injiciranje ..."  -> "Somatuline Autogel"

        Strategy: walk tokens left to right, stop at the first token that
        looks like a strength/number, a unit, or a known Slovenian dosage
        form/packaging keyword. Keep at most 4 leading tokens so long
        INN-style names (rokuronijev bromid hameln) survive but comma-joined
        pack descriptions don't leak in.
        """
        if not product:
            return ""

        # Slovenian text may include diacritics (c, s, z with caron) -- fold
        # to ASCII for matching against the stopword list, but keep the
        # original tokens for the returned name.
        tokens = product.replace(",", " ").split()
        name_tokens: list[str] = []

        for tok in tokens:
            folded = _fold_diacritics(tok.lower()).rstrip(".,/%")
            if re.match(r"^\d+([.,]\d+)?$", folded):
                break
            if re.match(r"^\d+([.,]\d+)?(mg|g|ml|mcg|iu|ie|%)/?.*$", folded):
                break
            if folded in _FORM_STOPWORDS:
                break
            name_tokens.append(tok)
            if len(name_tokens) >= 4:
                break

        result = " ".join(name_tokens).strip()
        return result if result else tokens[0]

    @staticmethod
    def _parse_date(raw: Any) -> str | None:
        """Parse a workbook cell (datetime, date, or string) to ISO-8601 date."""
        if raw is None:
            return None
        if isinstance(raw, datetime):
            return raw.date().isoformat()
        if isinstance(raw, date):
            return raw.isoformat()
        raw_str = str(raw).strip()
        if not raw_str or raw_str in ("-", "N/A", "null", "None"):
            return None
        try:
            from dateutil import parser as dtparser
            dt = dtparser.parse(raw_str, dayfirst=True)
            return dt.date().isoformat()
        except (ValueError, ImportError):
            pass
        return None


def _fold_diacritics(text: str) -> str:
    """Fold Slovenian diacritics (c/s/z with caron, etc.) to plain ASCII."""
    import unicodedata
    nfkd = unicodedata.normalize("NFKD", text)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


# -------------------------------------------------------------------------
# Standalone entrypoint
# -------------------------------------------------------------------------

if __name__ == "__main__":
    import json
    import os
    import sys

    from dotenv import load_dotenv

    load_dotenv()

    dry_run = os.environ.get("MEDERTI_DRY_RUN", "0").strip() == "1"

    if dry_run:
        from unittest.mock import MagicMock

        print("=" * 60)
        print("DRY RUN MODE  (MEDERTI_DRY_RUN=1)")
        print("Fetches live JAZMP data but makes NO database writes.")
        print("Set MEDERTI_DRY_RUN=0 to run against Supabase.")
        print("=" * 60)

        scraper = SloveniaJAZMPScraper(db_client=MagicMock())

        print("\n-- Fetching from JAZMP ...")
        raw = scraper.fetch()
        print(f"-- Raw records received : {len(raw)}")

        print("-- Normalising records ...")
        events = scraper.normalize(raw)
        print(f"-- Normalised events    : {len(events)}")

        if events:
            print("\n-- Sample event (first record, raw_record omitted):")
            sample = {k: v for k, v in events[0].items() if k != "raw_record"}
            print(json.dumps(sample, indent=2, default=str))

            from collections import Counter

            status_counts   = Counter(e["status"] for e in events)
            severity_counts = Counter(e.get("severity") for e in events)
            reason_counts   = Counter(e.get("reason_category") for e in events)

            print("\n-- Status breakdown:")
            for k, v in sorted(status_counts.items()):
                print(f"   {k:25s} {v}")
            print("\n-- Severity breakdown:")
            for k, v in sorted(severity_counts.items()):
                print(f"   {str(k):12s} {v}")
            print("\n-- Reason category breakdown:")
            for k, v in sorted(reason_counts.items()):
                print(f"   {str(k):30s} {v}")

        print("\n-- Dry run complete. No writes made to Supabase.")
        sys.exit(0)

    # -- Live run --
    print("=" * 60)
    print("LIVE RUN - writing to Supabase")
    print("=" * 60)

    scraper = SloveniaJAZMPScraper()
    summary = scraper.run()

    print("\n-- Scrape summary:")
    print(json.dumps(summary, indent=2, default=str))

    sys.exit(0 if summary["status"] in ("success", "duplicate") else 1)
