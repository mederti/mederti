"""
EMA — Withdrawn / Refused / Revoked Medicines Scraper
─────────────────────────────────────────────────────
Source:  European Medicines Agency — Medicines data download
Dataset: https://www.ema.europa.eu/en/documents/report/medicines-output-medicines-report_en.xlsx

Downloads the full medicines Excel (updated nightly by EMA), filters to
human medicines with status: Withdrawn, Refused, Revoked, Suspended, Lapsed.
Headers at row 9, data from row 10.

Source UUID:  10000000-0000-0000-0000-000000000028
Country code: EU
"""

from __future__ import annotations

import io
import re
from datetime import datetime, timezone

from backend.scrapers.base_recall_scraper import BaseRecallScraper


class EMARecallsScraper(BaseRecallScraper):
    """Scraper for EMA withdrawn authorisations and EU-level drug recalls."""

    SOURCE_ID:    str = "10000000-0000-0000-0000-000000000028"
    SOURCE_NAME:  str = "EMA — Withdrawn / Refused / Revoked Medicines"
    BASE_URL:     str = "https://www.ema.europa.eu/en/medicines/download-medicine-data"
    DATA_URL:     str = (
        "https://www.ema.europa.eu/en/documents/report/"
        "medicines-output-medicines-report_en.xlsx"
    )
    COUNTRY:      str = "European Union"
    COUNTRY_CODE: str = "EU"

    RATE_LIMIT_DELAY: float = 1.0

    # Statuses that count as recall-like events
    _RECALL_STATUSES: frozenset[str] = frozenset([
        "withdrawn", "refused", "revoked", "suspended", "lapsed", "expired",
    ])

    # ─────────────────────────────────────────────────────────────────────────
    # fetch()
    # ─────────────────────────────────────────────────────────────────────────

    def fetch(self) -> dict:
        """Download the EMA withdrawn authorisations Excel file."""
        self.log.info("Fetching EMA withdrawn authorisations", extra={"url": self.DATA_URL})

        try:
            resp = self._get(self.DATA_URL)
            self.log.info("EMA data fetched", extra={"bytes": len(resp.content)})
            return {"content": resp.content, "source": "xlsx", "url": self.DATA_URL}
        except Exception as exc:
            self.log.warning("EMA xlsx fetch failed", extra={"error": str(exc)})
            return {"content": None, "source": "none", "error": str(exc)}

    # ─────────────────────────────────────────────────────────────────────────
    # normalize()
    # ─────────────────────────────────────────────────────────────────────────

    def normalize(self, raw: dict | list) -> list[dict]:
        if isinstance(raw, dict) and raw.get("source") == "xlsx" and raw.get("content"):
            return self._parse_xlsx(raw["content"])

        self.log.warning("EMA: no data to normalise", extra={"raw": str(raw)[:100]})
        return []

    def _parse_xlsx(self, content: bytes) -> list[dict]:
        try:
            import openpyxl
        except ImportError:
            self.log.error("openpyxl not installed — run: pip install openpyxl")
            return []

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active

        # Find header row — it's the first row with 5+ non-None values
        header_row_idx = None
        headers: list[str] = []
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
            non_none = [v for v in row if v is not None]
            if len(non_none) >= 5:
                headers = [str(h or "").strip().lower() for h in row]
                header_row_idx = i
                break

        if not header_row_idx:
            self.log.warning("EMA: could not find header row")
            wb.close()
            return []

        self.log.info("EMA Excel headers", extra={
            "row": header_row_idx,
            "headers": headers[:15],
        })

        normalised: list[dict] = []
        total = 0
        skipped = 0
        today = datetime.now(timezone.utc).date().isoformat()

        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            total += 1
            try:
                rec = dict(zip(headers, row))

                # Filter: human medicines only
                category = str(rec.get("category", "") or "").strip().lower()
                if category and category != "human":
                    skipped += 1
                    continue

                # Filter: only recall-like statuses
                status = str(rec.get("medicine status", "") or "").strip().lower()
                if status not in self._RECALL_STATUSES:
                    skipped += 1
                    continue

                result = self._normalise_ema_record(rec, today)
                if result:
                    normalised.append(result)
                else:
                    skipped += 1
            except Exception as exc:
                skipped += 1
                self.log.debug("EMA row error", extra={"error": str(exc)})

        self.log.info("EMA normalisation done", extra={
            "total": total, "normalised": len(normalised), "skipped": skipped,
        })
        wb.close()
        return normalised

    def _normalise_ema_record(self, rec: dict, today: str) -> dict | None:
        # Brand name
        brand = str(rec.get("name of medicine", "") or "").strip()
        if not brand:
            return None

        # Active substance / INN = generic name
        inn = str(
            rec.get("international non-proprietary name (inn) / common name", "")
            or rec.get("active substance", "")
            or brand
        ).strip()

        # Medicine status → recall_type
        status = str(rec.get("medicine status", "") or "").strip().lower()
        recall_type_map = {
            "withdrawn": "market_withdrawal",
            "refused":   "refused",
            "revoked":   "revoked",
            "suspended": "suspended",
            "lapsed":    "lapsed",
            "expired":   "expired",
        }
        recall_type = recall_type_map.get(status, "market_withdrawal")

        # Date — try withdrawal date, then refusal date, then EC decision date
        date_raw = (
            rec.get("withdrawal / expiry / revocation / lapse of marketing authorisation date", "")
            or rec.get("refusal of marketing authorisation date", "")
            or rec.get("european commission decision date", "")
            or ""
        )
        announced_date = self._parse_date(str(date_raw)) or today

        # Therapeutic indication as reason
        indication = str(rec.get("therapeutic indication", "") or "").strip()
        reason = f"Marketing authorisation {status}" + (
            f". Indication: {indication[:200]}" if indication else ""
        )

        # MAH = manufacturer
        mah = rec.get("marketing authorisation developer / applicant / holder")
        mah_str = str(mah).strip()[:200] if mah else None

        # EMA product number as recall ref
        ema_num = str(rec.get("ema product number", "") or brand)[:80]

        # URL
        medicine_url = str(rec.get("medicine url", "") or self.BASE_URL).strip()

        return {
            "generic_name":     inn[:100],
            "brand_name":       brand if brand.lower() != inn.lower() else None,
            "manufacturer":     mah_str,
            "recall_class":     "Unclassified",
            "recall_type":      recall_type,
            "reason":           reason[:500] if reason else None,
            "reason_category":  "other",
            "lot_numbers":      [],
            "announced_date":   announced_date,
            "status":           "completed",
            "press_release_url": medicine_url,
            "confidence_score": 85,
            "recall_ref":       ema_num,
            "raw_record":       {k: str(v)[:200] for k, v in rec.items() if v is not None},
        }

    @staticmethod
    def _parse_date(raw: str) -> str | None:
        if not raw or raw in ("None", "nan", ""):
            return None
        # Excel often returns datetime objects as strings "YYYY-MM-DD HH:MM:SS"
        m = re.search(r"\d{4}-\d{2}-\d{2}", raw)
        if m:
            return m.group(0)
        for fmt in ("%d/%m/%Y", "%d.%m.%Y", "%B %d, %Y"):
            try:
                return datetime.strptime(raw[:20], fmt).date().isoformat()
            except Exception:
                pass
        return None

    @staticmethod
    def _map_reason(raw: str) -> str | None:
        if not raw:
            return "other"
        lower = raw.lower()
        if any(w in lower for w in ["contamination", "contaminated"]):
            return "contamination"
        if any(w in lower for w in ["label", "labelling"]):
            return "mislabelling"
        if any(w in lower for w in ["potency", "efficacy"]):
            return "subpotency"
        if "sterility" in lower:
            return "sterility"
        return "other"


# ─────────────────────────────────────────────────────────────────────────────
# Standalone entrypoint
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import json
    import os
    import sys

    from dotenv import load_dotenv

    load_dotenv()
    dry_run = os.environ.get("MEDERTI_DRY_RUN", "0").strip() == "1"

    if dry_run:
        from unittest.mock import MagicMock

        print("=" * 60)
        print("DRY RUN — EMA Recalls")
        print("=" * 60)
        scraper = EMARecallsScraper(db_client=MagicMock())
        raw = scraper.fetch()
        recalls = scraper.normalize(raw)
        print(f"── Normalised recalls: {len(recalls)}")
        if recalls:
            print(json.dumps({k: v for k, v in recalls[0].items() if k != "raw_record"}, indent=2, default=str))
        sys.exit(0)

    scraper = EMARecallsScraper()
    summary = scraper.run()
    print(json.dumps(summary, indent=2, default=str))
    sys.exit(0 if summary["status"] == "success" else 1)
