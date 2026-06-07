from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INK = "12211B"; GREEN = "1F6F4A"; LIGHT = "E6F0EA"; ZEBRA = "F4F8F6"
AMBER = "FFF3D6"; GREY = "5B6660"; WHITEF = "FFFFFF"
FONT = "Arial"

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, bold=True, color=INK, size=10)
        cell.fill = PatternFill("solid", fgColor=LIGHT)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = border

def body(ws, start_row, end_row, ncols, zebra=True):
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name=FONT, size=10, color=INK)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = border
            if zebra and (r - start_row) % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=ZEBRA)

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

STATUS_FILL = {
    "Live · yielding": "DCEEE3",
    "Live · 0 rows": AMBER,
    "Built · not wired": "F0DCDC",
}

wb = Workbook()

# ---------------- README ----------------
ws = wb.active; ws.title = "README"
ws.sheet_view.showGridLines = False
ws["A1"] = "Mederti — Source Inventory Matrix"
ws["A1"].font = Font(name=FONT, bold=True, size=20, color=GREEN)
ws["A2"] = "Companion to the Data Asset Inventory · for forward source mapping & prioritisation"
ws["A2"].font = Font(name=FONT, size=11, color=GREY)
notes = [
    ("", ""),
    ("Prepared for", "Chief Data Officer"),
    ("Generated", "3 June 2026"),
    ("Figures", "Live production counts queried 3 June 2026 (exact-count REST queries)."),
    ("", ""),
    ("How to read the status column", ""),
    ("Live · yielding", "Scraper wired in the daily cron AND producing rows in production."),
    ("Live · 0 rows", "Scraper wired in cron but returning zero production rows — likely pending deployment / in-flight. Fastest wins to chase."),
    ("Built · not wired", "Scraper code exists in the repo but is not scheduled. Engineering largely done; needs wiring."),
    ("", ""),
    ("Sheets", ""),
    ("1. Regulatory Feeds", "Primary shortage & recall sources, by agency — the core source matrix."),
    ("2. Reference & Supply", "Enrichment datasets and the supply-side / API-concentration sources."),
    ("3. Macro Catalogue", "The 134 catalogued macro / early-warning sources by signal category."),
    ("4. Coverage Summary", "Roll-ups: status tally, domain totals, regional depth."),
    ("", ""),
    ("Caveat", "Built-but-unshipped scrapers and pending migrations are excluded from 'yielding'. As deployments land, 'Live · 0 rows' rows convert to 'yielding'."),
]
r = 4
for k, v in notes:
    ws.cell(row=r, column=1, value=k).font = Font(name=FONT, bold=bool(k) and v != "" and not v[0].isdigit() and k not in ("Live · yielding","Live · 0 rows","Built · not wired","1. Regulatory Feeds","2. Reference & Supply","3. Macro Catalogue","4. Coverage Summary"), size=10, color=INK)
    ws.cell(row=r, column=1).font = Font(name=FONT, bold=True, size=10, color=INK)
    c2 = ws.cell(row=r, column=2, value=v)
    c2.font = Font(name=FONT, size=10, color=INK)
    c2.alignment = Alignment(wrap_text=True, vertical="top")
    if k in STATUS_FILL:
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=STATUS_FILL[k])
    r += 1
set_widths(ws, [26, 95])

# ---------------- Regulatory Feeds ----------------
ws = wb.create_sheet("Regulatory Feeds")
ws.sheet_view.showGridLines = False
hdr = ["Source / Agency", "Code", "Country", "ISO", "Region",
       "Shortage feed", "Recall feed", "Cadence", "Status",
       "Shortage rows (live)", "Recall rows (live)", "Notes"]
title = ws.cell(row=1, column=1, value="Regulatory Feeds — shortage & recall sources by agency")
title.font = Font(name=FONT, bold=True, size=13, color=GREEN)
ws.append([])  # row2 spacer handled below
for c, h in enumerate(hdr, 1):
    ws.cell(row=3, column=c, value=h)
style_header(ws, 3, len(hdr))

Y = "✓"; D = "—"
rows = [
 ["U.S. FDA","FDA","United States","US","North America",Y,Y,"Daily","Live · yielding",4469,17689,"Enforcement + MedWatch + Drugs@FDA"],
 ["Health Canada","HC","Canada","CA","North America",Y,Y,"Daily","Live · yielding",3233,4597,""],
 ["EMA","EMA","EU (supranational)","EU","Europe",Y,Y,"Daily","Live · yielding",101,369,"Centralised authorisations"],
 ["MHRA","MHRA","United Kingdom","GB","Europe",Y,Y,"Daily","Live · yielding",80,391,""],
 ["HPRA","HPRA","Ireland","IE","Europe",Y,D,"Daily","Live · yielding",778,None,""],
 ["BfArM","BfArM","Germany","DE","Europe",Y,D,"Daily","Live · yielding",740,None,"Recall scraper built, not wired"],
 ["ANSM","ANSM","France","FR","Europe",Y,Y,"Daily","Live · yielding",575,171,"Binary .xls export"],
 ["AIFA","AIFA","Italy","IT","Europe",Y,Y,"Daily","Live · yielding",4422,460,""],
 ["AEMPS","AEMPS","Spain","ES","Europe",Y,Y,"Daily","Live · yielding",3172,40,""],
 ["CBG-MEB","CBG-MEB","Netherlands","NL","Europe",Y,D,"Daily","Live · yielding",1010,None,""],
 ["Swissmedic","Swissmedic","Switzerland","CH","Europe",Y,D,"Daily","Live · yielding",5930,None,""],
 ["NoMA","NoMA","Norway","NO","Europe",Y,D,"Daily","Live · yielding",973,None,""],
 ["FIMEA","FIMEA","Finland","FI","Europe",Y,D,"Daily","Live · yielding",1815,None,""],
 ["EOF","EOF","Greece","GR","Europe",Y,D,"Daily","Live · yielding",440,None,""],
 ["FAMHP","FAMHP","Belgium","BE","Europe",Y,D,"Daily","Live · yielding",1014,None,""],
 ["INFARMED","INFARMED","Portugal","PT","Europe",Y,D,"Daily","Live · yielding",64,None,""],
 ["DKMA","DKMA","Denmark","DK","Europe",Y,D,"Daily","Live · 0 rows",0,None,"Cron-wired; no prod rows yet"],
 ["Läkemedelsverket","MPA","Sweden","SE","Europe",Y,D,"Daily","Live · 0 rows",0,None,"Cron-wired; no prod rows yet"],
 ["AGES","AGES","Austria","AT","Europe",Y,D,"Daily","Live · 0 rows",0,None,"Cron-wired; no prod rows yet"],
 ["OGYEI","OGYEI","Hungary","HU","Europe",Y,D,"Daily","Live · 0 rows",0,None,"Cron-wired; no prod rows yet"],
 ["SUKL","SUKL","Czechia","CZ","Europe",Y,D,"Daily","Live · 0 rows",0,None,"Cron-wired; no prod rows yet"],
 ["Ministry of Health","MZ","Poland","PL","Europe",Y,D,"Daily","Live · 0 rows",0,None,"Cron-wired; no prod rows yet"],
 ["TGA","TGA","Australia","AU","Asia-Pacific",Y,Y,"Daily","Live · yielding",1707,390,"Daily audit vs TGA MSI"],
 ["HSA","HSA","Singapore","SG","Asia-Pacific",Y,Y,"Daily","Live · yielding",321,0,"Recall scraper built, not wired"],
 ["Medsafe / PHARMAC","Medsafe","New Zealand","NZ","Asia-Pacific",Y,Y,"Daily","Live · yielding",290,186,""],
 ["PMDA","PMDA","Japan","JP","Asia-Pacific",Y,D,"Daily","Live · yielding",5672,None,""],
 ["NPRA","NPRA","Malaysia","MY","Asia-Pacific",Y,D,"Daily","Live · yielding",226,None,""],
 ["MFDS","MFDS","South Korea","KR","Asia-Pacific",Y,D,"Daily","Live · 0 rows",0,None,"Cron-wired; no prod rows yet"],
 ["ANVISA","ANVISA","Brazil","BR","Latin America",Y,D,"Daily","Live · 0 rows",0,None,"Cron-wired; no prod rows yet"],
 ["COFEPRIS","COFEPRIS","Mexico","MX","Latin America",Y,D,"Daily","Live · 0 rows",0,None,"Cron-wired; no prod rows yet"],
 ["ANMAT","ANMAT","Argentina","AR","Latin America",Y,D,"Daily","Live · 0 rows",0,None,"Cron-wired; no prod rows yet"],
 ["SAHPRA","SAHPRA","South Africa","ZA","Africa & Middle East",Y,D,"Daily","Live · 0 rows",0,None,"Cron-wired; no prod rows yet"],
 ["NAFDAC","NAFDAC","Nigeria","NG","Africa & Middle East",Y,D,"Daily","Live · 0 rows",0,None,"Cron-wired; no prod rows yet"],
 ["SFDA","SFDA","Saudi Arabia","SA","Africa & Middle East",Y,D,"Daily","Live · 0 rows",0,None,"Cron-wired; no prod rows yet"],
 ["MOHAP","MOHAP","United Arab Emirates","AE","Africa & Middle East",Y,D,"Daily","Live · yielding",20,None,""],
 ["NMPA","NMPA","China","CN","Asia-Pacific",Y,D,D,"Built · not wired",0,None,"High priority — code exists, not scheduled"],
 ["CDSCO","CDSCO","India","IN","Asia-Pacific",Y,D,D,"Built · not wired",0,None,"High priority — code exists, not scheduled"],
 ["Ministry of Health","MOH","Israel","IL","Africa & Middle East",Y,D,D,"Built · not wired",0,None,"Code exists, not scheduled"],
 ["TITCK","TITCK","Turkey","TR","Asia-Pacific",Y,D,D,"Built · not wired",0,None,"Code exists, not scheduled"],
 ["Drug Office","DO","Hong Kong","HK","Asia-Pacific",Y,D,D,"Built · not wired",0,None,"Code exists, not scheduled"],
]
start = 4
for row in rows:
    ws.append([None]*0)  # noop
for i, row in enumerate(rows):
    rr = start + i
    for c, val in enumerate(row, 1):
        ws.cell(row=rr, column=c, value=val)
end = start + len(rows) - 1
body(ws, start, end, len(hdr))
# numeric alignment + status fills
for rr in range(start, end + 1):
    for col in (4,6,7,10,11):
        ws.cell(row=rr, column=col).alignment = Alignment(horizontal="center", vertical="center")
    for col in (10,11):
        ws.cell(row=rr, column=col).number_format = '#,##0;-;"—"'
    st = ws.cell(row=rr, column=9).value
    if st in STATUS_FILL:
        ws.cell(row=rr, column=9).fill = PatternFill("solid", fgColor=STATUS_FILL[st])
        ws.cell(row=rr, column=9).font = Font(name=FONT, size=10, bold=True, color=INK)
ws.freeze_panes = "A4"
ws.auto_filter.ref = f"A3:L{end}"
set_widths(ws, [20, 11, 18, 6, 18, 11, 11, 9, 16, 14, 14, 34])

# ---------------- Reference & Supply ----------------
ws = wb.create_sheet("Reference & Supply")
ws.sheet_view.showGridLines = False
ws.cell(row=1, column=1, value="Reference, Enrichment & Supply-Side Sources").font = Font(name=FONT, bold=True, size=13, color=GREEN)
hdr2 = ["Dataset", "Type", "Provider", "Cadence", "Status", "Live rows", "Role / signal"]
for c, h in enumerate(hdr2, 1):
    ws.cell(row=3, column=c, value=h)
style_header(ws, 3, len(hdr2))
rs = [
 ["WHO ATC / DDD","Reference","WHO","On-demand","Live",None,"Therapeutic classification + defined daily dose"],
 ["RxNorm","Reference","US NLM","On-demand","Live",None,"US clinical drug nomenclature & ingredient mapping"],
 ["UNII","Reference","FDA","On-demand","Live",None,"Unique ingredient identifiers (salt-form resolution)"],
 ["SNOMED CT","Reference","SNOMED Intl","On-demand","Live",None,"Clinical terminology cross-walk"],
 ["EMA EPAR","Reference","EMA","On-demand","Live",None,"EU authorisation metadata"],
 ["OECD pharma","Reference","OECD","On-demand","Live",None,"Macro pharmaceutical market reference"],
 ["PharmaCompass","Supply ref","PharmaCompass","On-demand","Live",None,"API / supplier market reference"],
 ["FDA Drug Master File","Supply","FDA","Quarterly","Live · yielding",14009,"Active-ingredient maker signal → api_suppliers (14,009 rows)"],
 ["FDA DECRS","Supply","FDA","Quarterly","Live","enrich","Country-of-manufacture enrichment on api_suppliers"],
 ["WHO Prequalification","Supply","WHO","Quarterly","Live","badge","Prequalified global API makers ('WHO-PQ' badge)"],
 ["FDA inspections","Supply","FDA","Quarterly","Live",None,"Facility inspection / OAI signal"],
]
start = 4
for i, row in enumerate(rs):
    for c, val in enumerate(row, 1):
        ws.cell(row=start+i, column=c, value=val)
end = start + len(rs) - 1
body(ws, start, end, len(hdr2))
for rr in range(start, end+1):
    ws.cell(row=rr, column=6).alignment = Alignment(horizontal="center", vertical="center")
    v = ws.cell(row=rr, column=6).value
    if isinstance(v, int):
        ws.cell(row=rr, column=6).number_format = '#,##0'
set_widths(ws, [24, 12, 14, 12, 16, 12, 52])
ws.freeze_panes = "A4"

# ---------------- Macro Catalogue ----------------
ws = wb.create_sheet("Macro Catalogue")
ws.sheet_view.showGridLines = False
ws.cell(row=1, column=1, value="Macro Intelligence Catalogue — 134 sources by signal category").font = Font(name=FONT, bold=True, size=13, color=GREEN)
for c, h in enumerate(["Signal category", "Sources", "Example use in forward model"], 1):
    ws.cell(row=3, column=c, value=h)
style_header(ws, 3, 3)
cats = [
 ["Availability ground-truth",20,"Real-world stock-out confirmation"],
 ["Logistics",15,"Freight / port / cold-chain disruption"],
 ["Macro",12,"FX, inflation, broad market context"],
 ["Procurement",11,"Tender outcomes, GPO award shifts"],
 ["Early warning",10,"Leading shortage precursors"],
 ["External shocks",9,"Disasters, conflict, plant events"],
 ["Reference data",8,"Identity & classification backbone"],
 ["Pricing",8,"Price/tender movement (currently thin in DB)"],
 ["Pipeline",7,"Approvals & launches ahead of supply"],
 ["Data portals / discovery",7,"Source-discovery feeds"],
 ["Sanctions",6,"Trade-restriction exposure"],
 ["Trade",5,"Import/export flow signal"],
 ["Utilization",5,"Demand / consumption signal"],
 ["Corporate disclosure",5,"Manufacturer guidance & filings"],
 ["Public health",4,"Outbreak / demand-surge precursors"],
 ["Funding & aid flows",2,"Donor / aid procurement signal"],
]
start = 4
for i, row in enumerate(cats):
    for c, val in enumerate(row, 1):
        ws.cell(row=start+i, column=c, value=val)
end = start + len(cats) - 1
body(ws, start, end, 3)
for rr in range(start, end+1):
    ws.cell(row=rr, column=2).alignment = Alignment(horizontal="center", vertical="center")
# total row
tot = end + 1
ws.cell(row=tot, column=1, value="TOTAL").font = Font(name=FONT, bold=True, size=10, color=INK)
tc = ws.cell(row=tot, column=2, value=134)
tc.font = Font(name=FONT, bold=True, size=10, color=INK)
tc.alignment = Alignment(horizontal="center")
for c in range(1,4):
    ws.cell(row=tot, column=c).fill = PatternFill("solid", fgColor=LIGHT)
    ws.cell(row=tot, column=c).border = border
set_widths(ws, [28, 10, 46])
ws.freeze_panes = "A4"

# ---------------- Coverage Summary ----------------
ws = wb.create_sheet("Coverage Summary")
ws.sheet_view.showGridLines = False
ws.cell(row=1, column=1, value="Coverage Summary").font = Font(name=FONT, bold=True, size=14, color=GREEN)

ws.cell(row=3, column=1, value="Source status tally (Regulatory Feeds)").font = Font(name=FONT, bold=True, size=11, color=INK)
for c,h in enumerate(["Status","Count"],1):
    ws.cell(row=4, column=c, value=h)
style_header(ws,4,2)
stat_rows = [
 ["Live · yielding", 22],
 ["Live · 0 rows", 13],
 ["Built · not wired", 5],
 ["Total agencies", 40],
]
for i,(k,v) in enumerate(stat_rows):
    ws.cell(row=5+i, column=1, value=k)
    ws.cell(row=5+i, column=2, value=v)
body(ws,5,8,2)
for i in range(3):
    ws.cell(row=5+i, column=1).fill = PatternFill("solid", fgColor=STATUS_FILL[stat_rows[i][0]])
ws.cell(row=8,column=1).font = Font(name=FONT, bold=True, size=10)
ws.cell(row=8,column=2).font = Font(name=FONT, bold=True, size=10)
for rr in range(5,9):
    ws.cell(row=rr,column=2).alignment = Alignment(horizontal="center")

ws.cell(row=3, column=4, value="Live data volumes by domain").font = Font(name=FONT, bold=True, size=11, color=INK)
for c,h in zip((4,5),["Domain","Live rows"]):
    ws.cell(row=4, column=c, value=h)
style_header(ws,4,5)
# clear D4/E4 styling region overlap: re-style D4:E4
for c in (4,5):
    cell = ws.cell(row=4, column=c)
    cell.font = Font(name=FONT, bold=True, color=INK, size=10)
    cell.fill = PatternFill("solid", fgColor=LIGHT); cell.border = border
dom = [
 ["Shortage events",37065],["— active",28418],["— resolved",6439],["— anticipated",2204],
 ["— synthetic (recall-derived)",1589],
 ["Recall records",24293],["Recall→shortage links",78897],
 ["Canonical drugs",17835],["Brand/product catalogue",160977],
 ["API suppliers",14009],["Macro sources catalogued",134],["Registered data sources (41 active)",57],
]
for i,(k,v) in enumerate(dom):
    ws.cell(row=5+i, column=4, value=k)
    e = ws.cell(row=5+i, column=5, value=v); e.number_format='#,##0'; e.alignment=Alignment(horizontal="right")
de = 5+len(dom)-1
for r in range(5,de+1):
    for c in (4,5):
        cell=ws.cell(row=r,column=c); cell.border=border; cell.font=Font(name=FONT,size=10,color=INK)
        if (r-5)%2==1: cell.fill=PatternFill("solid",fgColor=ZEBRA)
    if ws.cell(row=r,column=4).value and str(ws.cell(row=r,column=4).value).startswith("—"):
        ws.cell(row=r,column=4).font=Font(name=FONT,size=10,italic=True,color=GREY)

note = ws.cell(row=de+2, column=4, value="Source: live Supabase REST exact-count queries, 3 June 2026.")
note.font = Font(name=FONT, size=9, italic=True, color=GREY)
set_widths(ws, [22, 10, 3, 30, 14])

wb.save("analysis/Mederti_Source_Inventory_Matrix.xlsx")
print("saved Mederti_Source_Inventory_Matrix.xlsx")
