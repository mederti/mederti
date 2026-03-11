"""
SFDA Saudi Arabia Drug Shortage Scraper
---------------------------------------
Source:  Saudi Food and Drug Authority - Drug Shortage List
URL:     https://www.sfda.gov.sa/en/currentlyInShortageList

The SFDA publishes a list of currently-in-shortage drugs on their website.
This scraper fetches the Excel export (richest data: 10 columns including
generic name, trade name, registration number, agent, manufacturer, reason,
start date, and duration). Falls back to the paginated JSON search API if
the Excel endpoint is unavailable.

Data source UUID:  10000000-0000-0000-0000-000000000043
Country:           Saudi Arabia
Country code:      SA
Confidence:        70/100 (self-reported by pharmaceutical agents)

Excel endpoint:  /GetExcel.php?ftype=CurrentlyInShortage
JSON API:        /GetCurrentlyInShortageSearch.php?AgentName=&ScientificName=&...

Cron:  Every 24 hours  (SFDA updates infrequently)
"""

from __future__ import annotations

import io
import re
from datetime import date, datetime, timezone
from typing import Any

from backend.scrapers.base_scraper import BaseScraper, ScraperError
from backend.utils.reason_mapper import map_reason_category


class SFDAScraper(BaseScraper):
    SOURCE_ID: str    = "10000000-0000-0000-0000-000000000043"
    SOURCE_NAME: str  = "Saudi Food and Drug Authority - Drug Shortage List"
    BASE_URL: str     = "https://www.sfda.gov.sa/en/currentlyInShortageList"
    COUNTRY: str      = "Saudi Arabia"
    COUNTRY_CODE: str = "SA"

    RATE_LIMIT_DELAY: float = 2.0
    REQUEST_TIMEOUT: float  = 60.0   # SFDA can be slow
    SCRAPER_VERSION: str    = "1.1.0"

    EXCEL_URL: str = "https://www.sfda.gov.sa/GetExcel.php?ftype=CurrentlyInShortage"
    JSON_API_URL: str = (
        "https://www.sfda.gov.sa/GetCurrentlyInShortageSearch.php"
        "?AgentName=&ScientificName=&RegistrationNo=&ShortageReason=&TradeName="
    )

    # Known SFDA shortage reasons -> reason_category
    _REASON_MAP: dict[str, str] = {
        "commercial /manufacturing issue": "manufacturing_issue",
        "commercial/manufacturing issue":  "manufacturing_issue",
        "commercial issue":                "manufacturing_issue",
        "manufacturing issue":             "manufacturing_issue",
        "global shortage":                 "supply_chain",
        "mah/agent changed":               "supply_chain",
        "mah changed":                     "supply_chain",
        "agent changed":                   "supply_chain",
        "regulations related issue":       "regulatory_action",
        "regulations related":             "regulatory_action",
        "other issue":                     "unknown",
        "other":                           "unknown",
    }

    # -------------------------------------------------------------------------
    # fetch()
    # -------------------------------------------------------------------------

    def fetch(self) -> list[dict]:
        """
        Fetch SFDA shortage data.

        Strategy:
        1. Try Excel download (richest data: 10 columns, ~1700 records).
        2. Fall back to paginated JSON search API.
        """
        self.log.info("Scrape started", extra={
            "source": self.SOURCE_NAME,
            "url": self.BASE_URL,
        })

        # Primary: Excel download
        try:
            return self._fetch_excel()
        except Exception as exc:
            self.log.warning(
                "Excel fetch failed, falling back to JSON API",
                extra={"error": str(exc)},
            )

        # Fallback: paginated JSON API
        return self._fetch_json_api()

    def _fetch_excel(self) -> list[dict]:
        """
        Download and parse the SFDA Excel export.

        Excel columns (as of Mar 2026):
            SCIENTIFICNAME_EN, TRADENAME_EN, REGISTRATION_NO,
            Manufacturer_Name, UPDATE_TIME, Agent,
            SHORTAGE_TYPE_EN, SHORTAGE_START_DATE, duration,
            SHORTAGE_REASON_EN
        """
        import openpyxl

        self.log.info("Fetching SFDA Excel export", extra={"url": self.EXCEL_URL})
        resp = self._get(self.EXCEL_URL)

        wb = openpyxl.load_workbook(
            io.BytesIO(resp.content),
            read_only=True,
            data_only=True,
        )
        ws = wb.active

        # Read headers from row 1
        headers: list[str] = []
        for cell in ws[1]:
            headers.append(str(cell.value or "").strip())

        self.log.info("SFDA Excel headers", extra={"headers": headers, "rows": ws.max_row})

        # Read data rows
        records: list[dict] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rec = {}
            for h, v in zip(headers, row):
                if h:
                    rec[h] = v
            # Skip completely empty rows
            if any(v for v in rec.values() if v is not None and str(v).strip()):
                records.append(rec)

        wb.close()

        self.log.info(
            "SFDA Excel fetch complete",
            extra={"records": len(records)},
        )
        return records

    def _fetch_json_api(self) -> list[dict]:
        """
        Fetch all pages from the SFDA JSON search API.

        The API returns paginated JSON with fields:
            scientificnamE_EN, shortagE_TYPE_EN, shortagE_TYPE_AR,
            tradeName_Count
        Plus pagination metadata: currentPage, pageCount, pageSize, rowCount
        """
        self.log.info("Fetching SFDA JSON API", extra={"url": self.JSON_API_URL})

        # First page to get pagination info
        first_page = self._get_json(self.JSON_API_URL)

        if isinstance(first_page, dict):
            page_count = first_page.get("pageCount", 1)
            # Extract results from first page (key varies)
            results_key = None
            for key in ("results", "data", "items", "rows"):
                if key in first_page and isinstance(first_page[key], list):
                    results_key = key
                    break

            if results_key:
                records = list(first_page[results_key])
            else:
                # The response might be the array directly wrapped with metadata
                records = [
                    {k: v for k, v in first_page.items()
                     if k not in ("currentPage", "pageCount", "pageSize",
                                  "rowCount", "firstRowOnPage", "lastRowOnPage")}
                ]
        elif isinstance(first_page, list):
            records = list(first_page)
            page_count = 1
        else:
            self.log.warning("Unexpected SFDA API response type")
            return []

        # Fetch remaining pages
        if page_count > 1:
            self.log.info(f"Fetching {page_count - 1} additional pages")
            for page in range(2, page_count + 1):
                try:
                    url = f"{self.JSON_API_URL}&page={page}"
                    page_data = self._get_json(url)
                    if isinstance(page_data, dict) and results_key and results_key in page_data:
                        records.extend(page_data[results_key])
                    elif isinstance(page_data, list):
                        records.extend(page_data)
                except Exception as exc:
                    self.log.warning(f"Failed to fetch page {page}", extra={"error": str(exc)})

        self.log.info(
            "SFDA JSON API fetch complete",
            extra={"records": len(records), "pages": page_count},
        )
        return records

    # -------------------------------------------------------------------------
    # normalize()
    # -------------------------------------------------------------------------

    def normalize(self, raw: list[dict]) -> list[dict]:
        """Normalize SFDA records into standard shortage event dicts."""
        self.log.info(
            "Normalising SFDA records",
            extra={"source": self.SOURCE_NAME, "raw_count": len(raw)},
        )

        normalised: list[dict] = []
        skipped = 0
        today = date.today().isoformat()

        for rec in raw:
            try:
                result = self._normalise_record(rec, today)
                if result is None:
                    skipped += 1
                    continue
                normalised.append(result)
            except Exception as exc:
                skipped += 1
                self.log.warning(
                    "Failed to normalise SFDA record",
                    extra={"error": str(exc), "rec": str(rec)[:200]},
                )

        self.log.info(
            "Normalisation done",
            extra={"total": len(raw), "normalised": len(normalised), "skipped": skipped},
        )
        return normalised

    def _normalise_record(self, rec: dict, today: str) -> dict | None:
        """Convert a single SFDA record to a normalised shortage event dict."""
        # -- Drug name extraction (Excel fields first, then JSON API fields) --
        generic_name = (
            rec.get("SCIENTIFICNAME_EN")
            or rec.get("scientificnamE_EN")
            or rec.get("Trade Name")
            or rec.get("Drug Name")
            or rec.get("Name")
            or ""
        )
        if isinstance(generic_name, str):
            generic_name = generic_name.strip()
        else:
            generic_name = str(generic_name).strip()

        if not generic_name:
            return None

        # -- Trade / brand name --
        trade_name = (
            rec.get("TRADENAME_EN")
            or rec.get("Trade Name")
            or ""
        )
        if isinstance(trade_name, str):
            trade_name = trade_name.strip()
        else:
            trade_name = str(trade_name).strip()

        brand_names = [trade_name] if trade_name and trade_name != generic_name else []

        # -- Shortage reason --
        raw_reason = (
            rec.get("SHORTAGE_REASON_EN")
            or rec.get("Shortage Reason")
            or rec.get("shortage_reason")
            or ""
        )
        if isinstance(raw_reason, str):
            raw_reason = raw_reason.strip()
        else:
            raw_reason = str(raw_reason).strip()

        reason_category = self._map_reason(raw_reason)

        # -- Start date (Excel has actual dates) --
        raw_start = rec.get("SHORTAGE_START_DATE")
        start_date = self._parse_date(raw_start) or today

        # -- Update date --
        raw_update = rec.get("UPDATE_TIME")
        update_date = self._parse_date(raw_update)

        # -- Registration number --
        reg_number = str(rec.get("REGISTRATION_NO") or "").strip()

        # -- Agent / manufacturer --
        agent = str(rec.get("Agent") or rec.get("Agent Name") or "").strip()
        manufacturer = str(rec.get("Manufacturer_Name") or rec.get("Manufacturer") or "").strip()

        # -- Duration in months --
        duration = rec.get("duration")
        if isinstance(duration, (int, float)) and duration > 0:
            duration_str = f"{int(duration)} months"
        else:
            duration_str = None

        # -- Status --
        shortage_type = str(
            rec.get("SHORTAGE_TYPE_EN")
            or rec.get("shortagE_TYPE_EN")
            or "Currently in Shortage"
        ).strip().lower()

        if "resolved" in shortage_type:
            status = "resolved"
        elif "anticipated" in shortage_type or "expected" in shortage_type:
            status = "anticipated"
        elif "discontinu" in shortage_type:
            status = "active"   # discontinuation = permanent shortage
        else:
            status = "active"

        # -- Build notes --
        notes_parts: list[str] = []
        if reg_number:
            notes_parts.append(f"Registration: {reg_number}")
        if agent:
            notes_parts.append(f"Agent: {agent}")
        if manufacturer:
            notes_parts.append(f"Manufacturer: {manufacturer}")
        if duration_str:
            notes_parts.append(f"Duration: {duration_str}")
        if update_date:
            notes_parts.append(f"Last updated: {update_date}")
        notes = "; ".join(notes_parts) or None

        return {
            "generic_name":              generic_name.title(),
            "brand_names":               brand_names,
            "status":                    status,
            "severity":                  "medium",
            "reason":                    raw_reason or None,
            "reason_category":           reason_category,
            "start_date":                start_date,
            "source_url":                self.BASE_URL,
            "notes":                     notes,
            "source_confidence_score":   70,
            "raw_record":                rec,
        }

    # -------------------------------------------------------------------------
    # Private helpers
    # -------------------------------------------------------------------------

    def _map_reason(self, raw: str) -> str:
        """Map SFDA reason string to canonical reason_category."""
        if not raw:
            return "unknown"
        lower = raw.strip().lower()
        for key, cat in self._REASON_MAP.items():
            if key in lower:
                return cat
        # Fallback to centralized mapper
        return map_reason_category(raw)

    @staticmethod
    def _parse_date(raw: Any) -> str | None:
        """Parse various SFDA date formats to ISO-8601 date string."""
        if raw is None:
            return None
        if isinstance(raw, datetime):
            return raw.date().isoformat()
        if isinstance(raw, date):
            return raw.isoformat()
        raw_str = str(raw).strip()
        if not raw_str or raw_str in ("-", "N/A", "null", "None"):
            return None
        try:
            from dateutil import parser as dtparser
            dt = dtparser.parse(raw_str)
            return dt.date().isoformat()
        except (ValueError, ImportError):
            pass
        return None

    @staticmethod
    def _extract_generic(trade_name: str) -> str:
        """
        Extract a generic-ish name from an SFDA trade name.

        Examples:
            "AUGMENTIN 1G TAB"          -> "Augmentin"
            "PANADOL EXTRA 500MG/65MG"  -> "Panadol Extra"
            "Metformin HCl 850mg"       -> "Metformin HCl"
        """
        if not trade_name:
            return trade_name

        # Strip common dosage suffixes: numbers followed by mg/g/ml/mcg/iu etc.
        name = re.split(
            r'\s+\d+[\.,]?\d*\s*(?:mg|g|ml|mcg|iu|%|tab|cap|inj|amp|vial)',
            trade_name,
            maxsplit=1,
            flags=re.IGNORECASE,
        )[0].strip()

        # Remove trailing dosage form words
        name = re.sub(
            r'\s+(tablets?|capsules?|injection|solution|suspension|syrup|cream|ointment'
            r'|drops?|inhaler|powder|patch|suppository|vials?|ampoules?)$',
            '',
            name,
            flags=re.IGNORECASE,
        ).strip()

        return name.title() if name else trade_name.title()


# -------------------------------------------------------------------------
# Standalone entrypoint
# -------------------------------------------------------------------------

if __name__ == "__main__":
    import json
    import os
    import sys

    from dotenv import load_dotenv

    load_dotenv()

    dry_run = os.environ.get("MEDERTI_DRY_RUN", "0").strip() == "1"

    if dry_run:
        from unittest.mock import MagicMock

        print("=" * 60)
        print("DRY RUN MODE  (MEDERTI_DRY_RUN=1)")
        print("Fetches live SFDA data but makes NO database writes.")
        print("Set MEDERTI_DRY_RUN=0 to run against Supabase.")
        print("=" * 60)

        scraper = SFDAScraper(db_client=MagicMock())

        print("\n-- Fetching from SFDA ...")
        raw = scraper.fetch()
        print(f"-- Raw records received : {len(raw)}")

        print("-- Normalising records ...")
        events = scraper.normalize(raw)
        print(f"-- Normalised events    : {len(events)}")

        if events:
            print("\n-- Sample event (first record, raw_record omitted):")
            sample = {k: v for k, v in events[0].items() if k != "raw_record"}
            print(json.dumps(sample, indent=2, default=str))

            from collections import Counter

            status_counts   = Counter(e["status"] for e in events)
            severity_counts = Counter(e.get("severity") for e in events)
            reason_counts   = Counter(e.get("reason_category") for e in events)

            print("\n-- Status breakdown:")
            for k, v in sorted(status_counts.items()):
                print(f"   {k:25s} {v}")
            print("\n-- Severity breakdown:")
            for k, v in sorted(severity_counts.items()):
                print(f"   {str(k):12s} {v}")
            print("\n-- Reason category breakdown:")
            for k, v in sorted(reason_counts.items()):
                print(f"   {str(k):30s} {v}")

        print("\n-- Dry run complete. No writes made to Supabase.")
        sys.exit(0)

    # -- Live run --
    print("=" * 60)
    print("LIVE RUN - writing to Supabase")
    print("=" * 60)

    scraper = SFDAScraper()
    summary = scraper.run()

    print("\n-- Scrape summary:")
    print(json.dumps(summary, indent=2, default=str))

    sys.exit(0 if summary["status"] in ("success", "duplicate") else 1)
