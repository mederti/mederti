"""
Portugal INFARMED Drug Shortage / Discontinuation Scraper
─────────────────────────────────────────────────────────
Source:  INFARMED — Gestão da Disponibilidade do Medicamento
URL:     https://www.infarmed.pt/web/infarmed/gestao-da-disponibilidade-do-medicamento

INFARMED (National Authority of Medicines and Health Products) publishes
information on drug supply disruptions and discontinuations in Portugal.
The page is in Portuguese. Data appears in two forms:
    1. Highlighted shortage narratives under "Situações de escassez em monitorização"
    2. A downloadable Excel file: the "Lista de exportação temporariamente suspensa"
       (export-suspended list) which contains structured drug shortage data.

Portuguese key terms:
    descontinuação temporária   = temporary discontinuation → active
    descontinuação definitiva   = permanent discontinuation → active (discontinued)
    reativação                  = reactivation              → resolved
    ruptura / rutura            = shortage / rupture         → active
    indisponível                = unavailable                → active
    disponível                  = available                  → resolved
    suspensão                   = suspension                 → active

Data source UUID:  10000000-0000-0000-0000-000000000048
Country:           Portugal
Country code:      PT
Confidence:        86/100 (official regulator, may require translation)

Cron:  Every 24 hours
"""

from __future__ import annotations

import io
import re
from datetime import date, datetime, timezone
from typing import Any
from urllib.parse import urljoin

from bs4 import BeautifulSoup

from backend.scrapers.base_scraper import BaseScraper, ScraperError
from backend.utils.reason_mapper import map_reason_category


class PortugalInfarmedScraper(BaseScraper):
    SOURCE_ID: str    = "10000000-0000-0000-0000-000000000048"
    SOURCE_NAME: str  = "INFARMED — Gestão da Disponibilidade do Medicamento"
    BASE_URL: str     = (
        "https://www.infarmed.pt/web/infarmed/"
        "gestao-da-disponibilidade-do-medicamento"
    )
    COUNTRY: str      = "Portugal"
    COUNTRY_CODE: str = "PT"

    RATE_LIMIT_DELAY: float = 2.0
    REQUEST_TIMEOUT: float  = 60.0
    SCRAPER_VERSION: str    = "2.0.0"

    # Portuguese status keywords -> standard status
    _STATUS_MAP: dict[str, str] = {
        "descontinuação temporária":    "active",
        "descontinuacao temporaria":    "active",
        "temporária":                   "active",
        "temporaria":                   "active",
        "descontinuação definitiva":    "active",
        "descontinuacao definitiva":    "active",
        "definitiva":                   "active",
        "ruptura":                      "active",
        "rutura":                       "active",
        "shortage":                     "active",
        "indisponível":                 "active",
        "indisponivel":                 "active",
        "unavailable":                  "active",
        "suspensão":                    "active",
        "suspensao":                    "active",
        "reativação":                   "resolved",
        "reativacao":                   "resolved",
        "reactivation":                 "resolved",
        "disponível":                   "resolved",
        "disponivel":                   "resolved",
        "available":                    "resolved",
        "resolved":                     "resolved",
        "resolvido":                    "resolved",
        "resolvida":                    "resolved",
    }

    # Portuguese reason keywords -> reason_category
    _REASON_MAP: dict[str, str] = {
        "fabricação":           "manufacturing_issue",
        "fabricacao":           "manufacturing_issue",
        "fabrico":             "manufacturing_issue",
        "produção":             "manufacturing_issue",
        "producao":             "manufacturing_issue",
        "manufacturing":        "manufacturing_issue",
        "qualidade":            "manufacturing_issue",
        "quality":              "manufacturing_issue",
        "matéria-prima":        "raw_material",
        "materia-prima":        "raw_material",
        "matéria prima":        "raw_material",
        "materia prima":        "raw_material",
        "raw material":         "raw_material",
        "substância ativa":     "raw_material",
        "substancia ativa":     "raw_material",
        "procura":              "demand_surge",
        "demanda":              "demand_surge",
        "demand":               "demand_surge",
        "cadeia de abastecimento": "supply_chain",
        "abastecimento":        "supply_chain",
        "supply":               "supply_chain",
        "logística":            "supply_chain",
        "logistica":            "supply_chain",
        "distribuição":         "distribution",
        "distribuicao":         "distribution",
        "distribution":         "distribution",
        "importação":           "distribution",
        "importacao":           "distribution",
        "descontinuação":       "discontinuation",
        "descontinuacao":       "discontinuation",
        "discontinu":           "discontinuation",
        "retirada":             "discontinuation",
        "withdrawal":           "discontinuation",
        "regulamentar":         "regulatory_action",
        "regulatory":           "regulatory_action",
        "autorização":          "regulatory_action",
        "autorizacao":          "regulatory_action",
        "comercial":            "supply_chain",
        "commercial":           "supply_chain",
    }

    # ─────────────────────────────────────────────────────────────────────────
    # fetch()
    # ─────────────────────────────────────────────────────────────────────────

    def fetch(self) -> list[dict]:
        """
        Fetch INFARMED shortage/discontinuation data.

        Strategy:
        1. GET the main page and parse highlighted shortage entries
           from the "Situações de escassez em monitorização" section.
        2. Look for the export-suspended list Excel file on the page,
           download and parse it for structured drug shortage data.
        3. Fall back to HTML table parsing if neither yields results.
        """
        self.log.info("Scrape started", extra={
            "source": self.SOURCE_NAME,
            "url": self.BASE_URL,
        })

        resp = self._get(self.BASE_URL)
        soup = BeautifulSoup(resp.text, "html.parser")

        records: list[dict] = []

        # Strategy 1: parse highlighted shortage narratives
        highlighted = self._parse_highlighted_shortages(soup)
        if highlighted:
            self.log.info(
                "INFARMED highlighted shortages parsed",
                extra={"records": len(highlighted)},
            )
            records.extend(highlighted)

        # Strategy 2: download and parse export-suspended Excel file
        excel_records = self._try_export_suspended_excel(soup)
        if excel_records:
            self.log.info(
                "INFARMED export-suspended Excel parsed",
                extra={"records": len(excel_records)},
            )
            records.extend(excel_records)

        # Strategy 3: fall back to HTML tables on the page
        if not records:
            table_records = self._parse_tables(soup)
            if table_records:
                self.log.info(
                    "INFARMED table parse complete",
                    extra={"records": len(table_records)},
                )
                records.extend(table_records)

        self.log.info(
            "INFARMED fetch complete",
            extra={"total_records": len(records)},
        )
        return records

    def _parse_highlighted_shortages(self, soup: BeautifulSoup) -> list[dict]:
        """
        Parse the "Situações de escassez em monitorização" section from
        the main INFARMED availability page.  This section contains
        narrative descriptions of drugs currently in shortage.
        """
        records: list[dict] = []

        # Use .journal-content-article first (the full-content container);
        # other selectors may match small fragments on this Liferay page.
        article = (
            soup.select_one(".journal-content-article")
            or soup.select_one("#content")
            or soup.select_one("main")
        )
        if not article:
            self.log.info("No article content found on INFARMED page")
            return records

        text = article.get_text(separator="\n", strip=True)
        lines = text.split("\n")

        # Find the "Situações de escassez em monitorização" section
        start_idx = None
        for i, line in enumerate(lines):
            if "escassez em monitorização" in line.lower():
                start_idx = i + 1
                break

        if start_idx is None:
            self.log.info("No 'Situações de escassez' section found")
            return records

        # Parse drug entries until we hit a section-end marker
        end_markers = {"ruturas", "faltas", "cessação de comercialização"}
        section_lines: list[str] = []
        for i in range(start_idx, len(lines)):
            if lines[i].strip().lower() in end_markers:
                break
            section_lines.append(lines[i])

        # Identify drug entry blocks: a heading followed by description lines
        current_heading: str | None = None
        current_desc_lines: list[str] = []

        for line in section_lines:
            stripped = line.strip()
            if not stripped:
                continue

            # A heading is a short line (drug name / category title) that
            # doesn't start with typical narrative markers or link text.
            # Must be at least 4 chars and not just a number/punctuation.
            is_heading = (
                4 <= len(stripped) < 80
                and not stripped.startswith((
                    "A ", "O ", "Os ", "Esta ", "Em ", "No ", "Sobre ",
                    "Consulte", "Circular", "EMA", "Orientações",
                    "Indisponibilidade", "Recomendações", "Dificuldade",
                    "EU ", "Shortage", "(", "ou ",
                ))
                and "poderá consultar" not in stripped.lower()
                and "http" not in stripped.lower()
                and not stripped.replace(".", "").replace(",", "").isdigit()
            )

            if is_heading and current_heading is not None:
                # Save previous entry
                desc = " ".join(current_desc_lines).strip()
                if desc:
                    records.append(self._build_highlighted_record(
                        current_heading, desc,
                    ))
                current_heading = stripped
                current_desc_lines = []
            elif is_heading and current_heading is None:
                current_heading = stripped
            else:
                current_desc_lines.append(stripped)

        # Don't forget the last entry
        if current_heading and current_desc_lines:
            desc = " ".join(current_desc_lines).strip()
            if desc:
                records.append(self._build_highlighted_record(
                    current_heading, desc,
                ))

        return records

    def _build_highlighted_record(self, heading: str, description: str) -> dict:
        """Build a record dict from a highlighted shortage heading + description."""
        # Map common headings to INN names
        heading_to_drugs: dict[str, list[str]] = {
            "agonistas do recetor da glp-1": ["semaglutide", "dulaglutide"],
            "pancreatina":                   ["pancreatin"],
            "quetiapina":                    ["quetiapine"],
            "estriol":                       ["estriol"],
            "metilfenidato":                 ["methylphenidate"],
            "sucralfato":                    ["sucralfate"],
        }

        heading_lower = heading.lower()
        drugs: list[str] | None = None
        for key, drug_list in heading_to_drugs.items():
            if key in heading_lower:
                drugs = drug_list
                break

        # Extract brand names from description
        brand_pattern = re.findall(
            r'(?:medicamento[s]?\s+)([A-Z][a-zà-ú]+(?:\s+[A-Z][a-zà-ú]*)*)',
            description,
        )
        brand_names = list(dict.fromkeys(
            b.strip() for b in brand_pattern if len(b) > 2
        ))

        # Determine reason from description text
        desc_lower = description.lower()
        reason: str | None = None
        if any(kw in desc_lower for kw in ("produção", "producao", "fabrico", "fabricação")):
            reason = "Manufacturing constraints"
        elif any(kw in desc_lower for kw in ("procura", "demanda", "prescrição")):
            reason = "Increased demand"
        elif "abastecimento" in desc_lower:
            reason = "Supply difficulty"

        return {
            "_source":          "highlighted",
            "heading":          heading,
            "description":      description[:500],
            "dci":              drugs[0] if drugs else heading,
            "brand_names":      brand_names[:5],
            "additional_drugs": drugs[1:] if drugs and len(drugs) > 1 else [],
            "reason":           reason,
        }

    def _try_export_suspended_excel(self, soup: BeautifulSoup) -> list[dict]:
        """
        Find and download the "Lista de exportação temporariamente suspensa"
        Excel file from the INFARMED page.  This is a structured list of
        drugs whose export is restricted due to shortages.
        """
        article = (
            soup.select_one(".journal-content-article")
            or soup.select_one("#content")
            or soup.select_one("main")
        )
        if not article:
            return []

        # Find the export-suspended list download link
        target_url: str | None = None
        for a in article.find_all("a", href=True):
            href = a["href"]
            text = a.get_text(strip=True).lower()
            if "/documents/" in href and (
                "exportação temporariamente suspensa" in text
                or "exportacao temporariamente suspensa" in text
                or ("lista em vigor" in text and "9423565" in href)
            ):
                target_url = href
                break

        if not target_url:
            # Fallback: any "lista em vigor" document link
            for a in article.find_all("a", href=True):
                href = a["href"]
                text = a.get_text(strip=True).lower()
                if "/documents/" in href and "lista em vigor" in text:
                    target_url = href
                    break

        if not target_url:
            self.log.info("No export-suspended Excel link found on INFARMED page")
            return []

        full_url = urljoin(self.BASE_URL, target_url)
        self.log.info(
            "Fetching INFARMED export-suspended Excel",
            extra={"url": full_url},
        )

        try:
            return self._parse_excel_download(full_url)
        except Exception as exc:
            self.log.warning(
                "Failed to parse INFARMED export-suspended Excel",
                extra={"url": full_url, "error": str(exc)},
            )
            return []

    def _parse_excel_download(self, url: str) -> list[dict]:
        """Download and parse an Excel file from INFARMED."""
        import openpyxl

        self.log.info("Fetching INFARMED Excel file", extra={"url": url})
        resp = self._get(url)

        wb = openpyxl.load_workbook(
            io.BytesIO(resp.content),
            read_only=True,
            data_only=True,
        )

        records: list[dict] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            self.log.info(
                "Processing INFARMED Excel sheet",
                extra={"sheet": sheet_name, "rows": ws.max_row},
            )

            # Read headers from row 1
            headers: list[str] = []
            for cell in ws[1]:
                headers.append(str(cell.value or "").strip().lower())

            self.log.info(
                "INFARMED Excel headers",
                extra={"sheet": sheet_name, "headers": headers},
            )

            # Read data rows
            for row in ws.iter_rows(min_row=2, values_only=True):
                rec: dict[str, Any] = {"_source": "excel", "_sheet": sheet_name}
                for h, v in zip(headers, row):
                    if h:
                        rec[h] = v
                if any(
                    v for v in rec.values()
                    if v is not None and str(v).strip()
                    and v not in ("excel", sheet_name)
                ):
                    records.append(rec)

        wb.close()

        self.log.info(
            "INFARMED Excel fetch complete",
            extra={"records": len(records)},
        )
        return records

    def _parse_tables(self, soup: BeautifulSoup) -> list[dict]:
        """Extract records from HTML <table> elements on the INFARMED page."""
        records: list[dict] = []

        tables = soup.find_all("table")
        for table in tables:
            # Read headers
            headers: list[str] = []
            header_row = table.find("thead")
            if header_row:
                for th in header_row.find_all(["th", "td"]):
                    headers.append(th.get_text(strip=True).lower())
            else:
                first_row = table.find("tr")
                if first_row:
                    for cell in first_row.find_all(["th", "td"]):
                        headers.append(cell.get_text(strip=True).lower())

            if not headers:
                continue

            # Check if this table looks like shortage/discontinuation data
            has_drug_col = any(
                kw in h
                for h in headers
                for kw in ("nome", "name", "dci", "inn", "medicamento",
                           "denominação", "denominacao", "substância",
                           "substancia", "produto", "product")
            )
            if not has_drug_col:
                continue

            # Parse data rows
            tbody = table.find("tbody") or table
            for tr in tbody.find_all("tr"):
                cells = tr.find_all(["td", "th"])
                if len(cells) < 2:
                    continue

                row: dict[str, str] = {"_source": "table"}
                for i, cell in enumerate(cells):
                    key = headers[i] if i < len(headers) else f"col_{i}"
                    row[key] = cell.get_text(strip=True)

                # Skip header-like rows
                if any(
                    row.get(h, "").lower() == h
                    for h in headers
                    if h
                ):
                    continue

                if any(v.strip() for v in row.values() if isinstance(v, str)):
                    records.append(row)

        return records

    # ─────────────────────────────────────────────────────────────────────────
    # normalize()
    # ─────────────────────────────────────────────────────────────────────────

    def normalize(self, raw: list[dict]) -> list[dict]:
        """Normalize INFARMED records into standard shortage event dicts."""
        self.log.info(
            "Normalising INFARMED records",
            extra={"source": self.SOURCE_NAME, "raw_count": len(raw)},
        )

        normalised: list[dict] = []
        skipped = 0
        today = date.today().isoformat()
        seen_drugs: set[str] = set()

        for rec in raw:
            try:
                source = rec.get("_source", "unknown")
                if source == "highlighted":
                    results = self._normalise_highlighted(rec, today, seen_drugs)
                elif source == "excel":
                    results = self._normalise_excel(rec, today, seen_drugs)
                else:
                    results = self._normalise_legacy(rec, today, seen_drugs)

                if not results:
                    skipped += 1
                    continue
                normalised.extend(results)
            except Exception as exc:
                skipped += 1
                self.log.warning(
                    "Failed to normalise INFARMED record",
                    extra={"error": str(exc), "rec": str(rec)[:200]},
                )

        self.log.info(
            "Normalisation done",
            extra={
                "total": len(raw),
                "normalised": len(normalised),
                "skipped": skipped,
            },
        )
        return normalised

    def _normalise_highlighted(
        self, rec: dict, today: str, seen: set[str],
    ) -> list[dict]:
        """Normalise a highlighted shortage entry (narrative section)."""
        results: list[dict] = []

        dci = rec.get("dci", "")
        if not dci:
            return []

        # Build list of drug names to create events for
        drug_names = [dci]
        for extra in rec.get("additional_drugs", []):
            if extra and extra != dci:
                drug_names.append(extra)

        for drug_name in drug_names:
            key = drug_name.strip().lower()
            if key in seen:
                continue
            seen.add(key)

            brand_names = rec.get("brand_names", [])
            raw_reason = rec.get("reason", "")
            reason_category = (
                self._map_reason(raw_reason) if raw_reason else "unknown"
            )

            # Build notes from description
            desc = rec.get("description", "")
            heading = rec.get("heading", "")
            notes_parts: list[str] = []
            if heading:
                notes_parts.append(f"Category: {heading}")
            if desc:
                notes_parts.append(f"Details: {desc[:200]}")

            results.append({
                "generic_name":              drug_name.strip().title(),
                "brand_names":               brand_names,
                "status":                    "active",
                "severity":                  "medium",
                "reason":                    raw_reason or None,
                "reason_category":           reason_category,
                "start_date":                today,
                "end_date":                  None,
                "estimated_resolution_date": None,
                "source_url":                self.BASE_URL,
                "notes":                     "; ".join(notes_parts) or None,
                "source_confidence_score":   86,
                "raw_record":                rec,
            })

        return results

    def _normalise_excel(
        self, rec: dict, today: str, seen: set[str],
    ) -> list[dict]:
        """Normalise an Excel row from the export-suspended list."""
        # Column names may vary between sheets; try common patterns
        generic_name = str(
            rec.get("dci/substância ativa")
            or rec.get("dci/substancia ativa")
            or rec.get("dci")
            or rec.get("substância ativa")
            or rec.get("substancia ativa")
            or ""
        ).strip()

        if not generic_name:
            return []

        key = generic_name.lower()
        if key in seen:
            return []
        seen.add(key)

        brand_name = str(
            rec.get("nome comercial") or rec.get("nome") or ""
        ).strip()
        brand_names = (
            [brand_name]
            if brand_name and brand_name.lower() != key
            else []
        )

        dosage = str(rec.get("dosagem") or rec.get("dose") or "").strip()
        form = str(
            rec.get("forma farmacêutica")
            or rec.get("forma farmaceutica")
            or rec.get("forma")
            or ""
        ).strip()
        holder = str(
            rec.get("titular de aim") or rec.get("titular") or ""
        ).strip()
        reg_num = str(
            rec.get("número de registo")
            or rec.get("numero de registo")
            or ""
        ).strip()
        desc_cits = str(
            rec.get("descrição cits") or rec.get("descricao cits") or ""
        ).strip()
        sheet = rec.get("_sheet", "")

        notes_parts: list[str] = []
        if holder:
            notes_parts.append(f"Holder: {holder}")
        if dosage:
            notes_parts.append(f"Dosage: {dosage}")
        if form:
            notes_parts.append(f"Form: {form}")
        if desc_cits:
            notes_parts.append(f"Presentation: {desc_cits}")
        if reg_num:
            notes_parts.append(f"Reg: {reg_num}")
        if sheet:
            notes_parts.append(f"Source: Export-suspended list ({sheet})")

        return [{
            "generic_name":              generic_name.strip().title(),
            "brand_names":               brand_names,
            "status":                    "active",
            "severity":                  "medium",
            "reason":                    "Export restricted due to shortage",
            "reason_category":           "supply_chain",
            "start_date":                today,
            "end_date":                  None,
            "estimated_resolution_date": None,
            "source_url":                self.BASE_URL,
            "notes":                     "; ".join(notes_parts) or None,
            "source_confidence_score":   86,
            "raw_record":                rec,
        }]

    def _normalise_legacy(
        self, rec: dict, today: str, seen: set[str],
    ) -> list[dict]:
        """Normalise a record from table parsing (legacy fallback)."""
        result = self._normalise_record(rec, today)
        if result is None:
            return []
        key = result["generic_name"].lower()
        if key in seen:
            return []
        seen.add(key)
        return [result]

    def _normalise_record(self, rec: dict, today: str) -> dict | None:
        """Convert a single INFARMED record to a normalised shortage event."""
        # -- Drug name extraction --
        generic_name = (
            rec.get("dci")
            or rec.get("inn")
            or rec.get("denominação comum internacional")
            or rec.get("denominacao comum internacional")
            or rec.get("substância ativa")
            or rec.get("substancia ativa")
            or rec.get("nome")
            or rec.get("name")
            or rec.get("medicamento")
            or rec.get("denominação")
            or rec.get("denominacao")
            or rec.get("produto")
            or rec.get("product")
            or ""
        )
        if isinstance(generic_name, str):
            generic_name = generic_name.strip()
        else:
            generic_name = str(generic_name).strip()

        # If no structured name, try raw_text extraction
        if not generic_name and rec.get("raw_text"):
            raw_text = rec["raw_text"]
            parts = re.split(
                r'\s*[-–|:]\s*(?=descontinua|ruptura|rutura|indisponível|indisponivel)',
                raw_text,
                maxsplit=1,
                flags=re.IGNORECASE,
            )
            generic_name = parts[0].strip()[:100]

        if not generic_name:
            return None

        # -- Brand / trade name --
        brand_name = (
            rec.get("nome comercial")
            or rec.get("trade name")
            or rec.get("marca")
            or rec.get("brand")
            or rec.get("nome do medicamento")
            or ""
        )
        if isinstance(brand_name, str):
            brand_name = brand_name.strip()
        else:
            brand_name = str(brand_name).strip()

        brand_names = (
            [brand_name]
            if brand_name and brand_name != generic_name
            else []
        )

        # -- Status --
        raw_status = str(
            rec.get("estado")
            or rec.get("status")
            or rec.get("situação")
            or rec.get("situacao")
            or rec.get("tipo")
            or rec.get("type")
            or ""
        ).strip().lower()

        if not raw_status and rec.get("raw_text"):
            raw_status = rec["raw_text"].lower()

        status = "active"
        for key, mapped_status in self._STATUS_MAP.items():
            if key in raw_status:
                status = mapped_status
                break

        # -- Reason --
        raw_reason = str(
            rec.get("motivo")
            or rec.get("reason")
            or rec.get("causa")
            or rec.get("justificação")
            or rec.get("justificacao")
            or ""
        ).strip()

        reason_category = self._map_reason(raw_reason)

        if reason_category == "unknown" and any(
            kw in raw_status
            for kw in ("descontinua", "definitiva", "withdrawal", "retirada")
        ):
            reason_category = "discontinuation"

        # -- Dates --
        raw_start = (
            rec.get("data início")
            or rec.get("data inicio")
            or rec.get("data de início")
            or rec.get("data de inicio")
            or rec.get("start date")
            or rec.get("data")
            or rec.get("date")
            or rec.get("data de notificação")
            or rec.get("data de notificacao")
            or ""
        )
        start_date = self._parse_date(raw_start) or today

        raw_end = (
            rec.get("data fim")
            or rec.get("data de fim")
            or rec.get("data prevista")
            or rec.get("data de resolução")
            or rec.get("data de resolucao")
            or rec.get("data reativação")
            or rec.get("data reativacao")
            or rec.get("end date")
            or rec.get("estimated end")
            or ""
        )
        end_date = self._parse_date(raw_end) if status == "resolved" else None
        estimated_resolution = (
            self._parse_date(raw_end) if status == "active" else None
        )

        # -- Dosage / presentation --
        dosage = str(
            rec.get("dosagem")
            or rec.get("dose")
            or rec.get("apresentação")
            or rec.get("apresentacao")
            or rec.get("forma farmacêutica")
            or rec.get("forma farmaceutica")
            or ""
        ).strip()

        # -- Holder / manufacturer --
        holder = str(
            rec.get("titular de aim")
            or rec.get("titular")
            or rec.get("fabricante")
            or rec.get("manufacturer")
            or rec.get("empresa")
            or rec.get("company")
            or ""
        ).strip()

        # -- Notes --
        notes_parts: list[str] = []
        if holder:
            notes_parts.append(f"Holder: {holder}")
        if dosage:
            notes_parts.append(f"Presentation: {dosage}")

        aim_num = str(
            rec.get("nº aim")
            or rec.get("n aim")
            or rec.get("aim")
            or rec.get("registration")
            or ""
        ).strip()
        if aim_num:
            notes_parts.append(f"AIM: {aim_num}")

        if raw_status and raw_status not in ("", "active", "resolved"):
            notes_parts.append(f"Type: {raw_status.title()}")

        if rec.get("raw_text") and not rec.get("dci") and not rec.get("nome"):
            notes_parts.append(f"Source text: {rec['raw_text'][:150]}")

        notes = "; ".join(notes_parts) or None

        return {
            "generic_name":              generic_name.title(),
            "brand_names":               brand_names,
            "status":                    status,
            "severity":                  "medium",
            "reason":                    raw_reason or None,
            "reason_category":           reason_category,
            "start_date":                start_date,
            "end_date":                  end_date,
            "estimated_resolution_date": estimated_resolution,
            "source_url":                self.BASE_URL,
            "notes":                     notes,
            "source_confidence_score":   86,
            "raw_record":                rec,
        }

    # ─────────────────────────────────────────────────────────────────────────
    # Private helpers
    # ─────────────────────────────────────────────────────────────────────────

    def _map_reason(self, raw: str) -> str:
        """Map INFARMED Portuguese reason string to canonical reason_category."""
        if not raw:
            return "unknown"
        lower = raw.strip().lower()
        for key, cat in self._REASON_MAP.items():
            if key in lower:
                return cat
        # Fallback to centralized mapper (handles English/French/etc.)
        return map_reason_category(raw)

    @staticmethod
    def _parse_date(raw: Any) -> str | None:
        """Parse various INFARMED date formats to ISO-8601 date string."""
        if raw is None:
            return None
        if isinstance(raw, datetime):
            return raw.date().isoformat()
        if isinstance(raw, date) and not isinstance(raw, datetime):
            return raw.isoformat()
        raw_str = str(raw).strip()
        if not raw_str or raw_str in ("-", "N/A", "null", "None", "", "n/a"):
            return None

        # Try ISO format first (YYYY-MM-DD or YYYY-MM-DDTHH:MM:SS)
        if "T" in raw_str:
            raw_str = raw_str[:10]
        iso_match = re.match(r'^(\d{4})-(\d{2})-(\d{2})$', raw_str)
        if iso_match:
            return raw_str

        # Portuguese date format: DD/MM/YYYY, DD-MM-YYYY, DD.MM.YYYY
        eu_match = re.match(
            r'^(\d{1,2})[/.\-](\d{1,2})[/.\-](\d{2,4})$', raw_str
        )
        if eu_match:
            day, month, year = eu_match.groups()
            if len(year) == 2:
                year = f"20{year}"
            try:
                dt = datetime(int(year), int(month), int(day))
                return dt.date().isoformat()
            except ValueError:
                pass

        # Portuguese month names (e.g., "15 de março de 2026")
        pt_months: dict[str, int] = {
            "janeiro": 1, "fevereiro": 2, "março": 3, "marco": 3,
            "abril": 4, "maio": 5, "junho": 6, "julho": 7,
            "agosto": 8, "setembro": 9, "outubro": 10,
            "novembro": 11, "dezembro": 12,
        }
        pt_match = re.match(
            r'(\d{1,2})\s+(?:de\s+)?(\w+)\s+(?:de\s+)?(\d{4})',
            raw_str.lower(),
        )
        if pt_match:
            day_str, month_str, year_str = pt_match.groups()
            month_num = pt_months.get(month_str)
            if month_num:
                try:
                    dt = datetime(int(year_str), month_num, int(day_str))
                    return dt.date().isoformat()
                except ValueError:
                    pass

        # Fallback: dateutil parser
        try:
            from dateutil import parser as dtparser
            dt = dtparser.parse(raw_str, dayfirst=True)
            return dt.date().isoformat()
        except (ValueError, ImportError):
            pass

        return None


# ─────────────────────────────────────────────────────────────────────────────
# Standalone entrypoint
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import json
    import os
    import sys

    from dotenv import load_dotenv

    load_dotenv()

    dry_run = os.environ.get("MEDERTI_DRY_RUN", "0").strip() == "1"

    if dry_run:
        from unittest.mock import MagicMock

        print("=" * 60)
        print("DRY RUN MODE  (MEDERTI_DRY_RUN=1)")
        print("Fetches live INFARMED data but makes NO database writes.")
        print("Set MEDERTI_DRY_RUN=0 to run against Supabase.")
        print("=" * 60)

        scraper = PortugalInfarmedScraper(db_client=MagicMock())

        print("\n-- Fetching from INFARMED ...")
        raw = scraper.fetch()
        print(f"-- Raw records received : {len(raw)}")

        print("-- Normalising records ...")
        events = scraper.normalize(raw)
        print(f"-- Normalised events    : {len(events)}")

        if events:
            print("\n-- Sample event (first record, raw_record omitted):")
            sample = {k: v for k, v in events[0].items() if k != "raw_record"}
            print(json.dumps(sample, indent=2, default=str))

            from collections import Counter

            status_counts   = Counter(e["status"] for e in events)
            severity_counts = Counter(e.get("severity") for e in events)
            reason_counts   = Counter(e.get("reason_category") for e in events)

            print("\n-- Status breakdown:")
            for k, v in sorted(status_counts.items()):
                print(f"   {k:25s} {v}")
            print("\n-- Severity breakdown:")
            for k, v in sorted(severity_counts.items()):
                print(f"   {str(k):12s} {v}")
            print("\n-- Reason category breakdown:")
            for k, v in sorted(reason_counts.items()):
                print(f"   {str(k):30s} {v}")

        print("\n-- Dry run complete. No writes made to Supabase.")
        sys.exit(0)

    # -- Live run --
    print("=" * 60)
    print("LIVE RUN - writing to Supabase")
    print("=" * 60)

    scraper = PortugalInfarmedScraper()
    summary = scraper.run()

    print("\n-- Scrape summary:")
    print(json.dumps(summary, indent=2, default=str))

    sys.exit(0 if summary["status"] in ("success", "duplicate") else 1)
