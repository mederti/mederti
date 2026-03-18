#!/usr/bin/env python3
"""
Japan MHLW Drug Supply Status Importer
───────────────────────────────────────
Source: https://www.mhlw.go.jp/stf/seisakunitsuite/bunya/kenkou_iryou/iryou/kouhatu-iyaku/04_00003.html
Data:   Daily Excel file with ~17,000 drug products and their supply status.

Writes to:
  - drug_catalogue (all 17k rows — JP drug registry)
  - shortage_events (rows with limited/suspended shipment status)

Column mapping (1-indexed from Excel):
  Col 3  成分名           → generic_name
  Col 4  規格単位         → strength
  Col 5  YJコード         → registration_number (YJ code)
  Col 6  品名             → brand_name
  Col 7  製造販売業者名   → sponsor (manufacturer)
  Col 8  製品区分         → product_category
  Col 12 出荷対応の状況   → shipment_status → availability_status
  Col 14 出荷停止等の理由 → reason
  Col 15 解消見込み       → resolution_prospect
  Col 16 解消見込み時期   → estimated_resolution_date
  Col 20 更新した日       → updated_at
"""
from __future__ import annotations

import hashlib
import logging
import os
import re
import sys
import tempfile
from datetime import datetime, timezone

import httpx
import openpyxl
from supabase import create_client

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_KEY = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

SOURCE_NAME = "MHLW"
SOURCE_COUNTRY = "JP"
DATA_SOURCE_ID = "10000000-0000-0000-0000-000000000037"  # PMDA/Japan

INDEX_URL = "https://www.mhlw.go.jp/stf/seisakunitsuite/bunya/kenkou_iryou/iryou/kouhatu-iyaku/04_00003.html"

# ── Status mapping ──────────────────────────────────────────────────
STATUS_MAP = {
    "①通常出荷":                       ("normal", "available"),
    "②限定出荷（自社の事情）":          ("limited_own", "limited"),
    "②限定出荷（他社品の影響）":        ("limited_other", "limited"),
    "②限定出荷（その他）":             ("limited_misc", "limited"),
    "③供給停止":                       ("suspended", "unavailable"),
    "④薬価削除品目（在庫消尽まで）":    ("delisted", "unavailable"),
    "⑤販売中止届提出済":               ("discontinued", "unavailable"),
}

SEVERITY_MAP = {
    "suspended":    "critical",
    "discontinued": "critical",
    "delisted":     "high",
    "limited_own":  "high",
    "limited_other":"medium",
    "limited_misc": "medium",
    "normal":       None,  # no shortage
}

REASON_MAP = {
    "１．品質問題":     "quality_issue",
    "２．製造上の問題":  "manufacturing_issue",
    "３．需要増":       "demand_surge",
    "４．原薬等の調達":  "raw_material_shortage",
    "５．GMP関連":      "regulatory_action",
    "６．販売中止":     "discontinuation",
    "７．ー":           None,
    "８．その他の理由":  "other",
}


def find_excel_url() -> str:
    """Scrape the MHLW index page to find the latest Excel download URL."""
    log.info("Fetching MHLW index page...")
    resp = httpx.get(INDEX_URL, follow_redirects=True, timeout=30)
    resp.raise_for_status()

    # Look for xlsx link
    match = re.search(r'href="(/content/[^"]+\.xlsx)"', resp.text)
    if match:
        return f"https://www.mhlw.go.jp{match.group(1)}"

    # Fallback: try known pattern (date-based filename)
    today = datetime.now().strftime("%y%m%d")
    return f"https://www.mhlw.go.jp/content/10800000/{today}iyakuhinkyoukyu.xlsx"


def download_excel(url: str) -> str:
    """Download Excel to temp file."""
    log.info(f"Downloading {url}...")
    resp = httpx.get(url, follow_redirects=True, timeout=60)
    resp.raise_for_status()
    path = os.path.join(tempfile.gettempdir(), "mhlw_shortage.xlsx")
    with open(path, "wb") as f:
        f.write(resp.content)
    log.info(f"Downloaded {len(resp.content):,} bytes → {path}")
    return path


def parse_excel(path: str) -> list[dict]:
    """Parse MHLW Excel into structured records."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]

    records = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:  # skip header rows
            continue

        cells = list(row)
        if len(cells) < 17 or not cells[2]:
            continue

        generic_name = str(cells[2] or "").strip()
        if not generic_name or len(generic_name) < 2:
            continue

        strength = str(cells[3] or "").strip()
        yj_code = str(cells[4] or "").strip()
        brand_name = str(cells[5] or "").strip()
        manufacturer = str(cells[6] or "").strip()
        product_category = str(cells[7] or "").strip()
        shipment_raw = str(cells[11] or "").strip()
        reason_raw = str(cells[13] or "").strip()
        resolution_prospect = str(cells[14] or "").strip()
        resolution_date_raw = cells[15]
        updated_raw = cells[19]

        # Map status
        status_key, avail_status = STATUS_MAP.get(shipment_raw, ("unknown", "unknown"))

        # Map reason category
        reason_cat = None
        for prefix, cat in REASON_MAP.items():
            if reason_raw.startswith(prefix):
                reason_cat = cat
                break

        # Parse resolution date
        resolution_date = None
        if resolution_date_raw:
            try:
                if isinstance(resolution_date_raw, datetime):
                    resolution_date = resolution_date_raw.strftime("%Y-%m-%d")
                else:
                    d = str(resolution_date_raw).strip()
                    if re.match(r"\d{4}-\d{2}-\d{2}", d):
                        resolution_date = d[:10]
            except Exception:
                pass

        # Parse updated date
        updated_at = None
        if updated_raw:
            try:
                if isinstance(updated_raw, datetime):
                    updated_at = updated_raw.isoformat()
                else:
                    d = str(updated_raw).strip()[:10]
                    if re.match(r"\d{4}-\d{2}-\d{2}", d):
                        updated_at = d
            except Exception:
                pass

        records.append({
            "generic_name": generic_name,
            "brand_name": brand_name,
            "strength": strength,
            "yj_code": yj_code,
            "manufacturer": manufacturer,
            "product_category": product_category,
            "status_key": status_key,
            "availability_status": avail_status,
            "severity": SEVERITY_MAP.get(status_key),
            "reason_raw": reason_raw,
            "reason_category": reason_cat,
            "resolution_prospect": resolution_prospect,
            "resolution_date": resolution_date,
            "updated_at": updated_at,
        })

    log.info(f"Parsed {len(records)} records from Excel")
    return records


def build_drug_index() -> dict[str, str]:
    """Build lookup from normalised generic name → drug ID."""
    log.info("Fetching drugs for linking...")
    index: dict[str, str] = {}
    offset = 0
    while True:
        batch = supabase.table("drugs").select("id, generic_name").range(offset, offset + 999).execute()
        for d in batch.data:
            if d.get("generic_name"):
                key = d["generic_name"].lower().strip()
                index[key] = d["id"]
        if len(batch.data) < 1000:
            break
        offset += 1000
    log.info(f"Drug index: {len(index)} entries")
    return index


def import_catalogue(records: list[dict], drug_index: dict[str, str]):
    """Insert all records into drug_catalogue."""
    # Check existing
    existing = supabase.table("drug_catalogue").select("id", count="exact").eq("source_name", SOURCE_NAME).execute()
    if existing.count and existing.count > 0:
        log.info(f"Already {existing.count} MHLW records in drug_catalogue — skipping catalogue insert")
        return

    rows = []
    linked = 0
    for r in records:
        # Try to link to drugs table
        drug_id = None
        name_lower = r["generic_name"].lower().strip()
        # Japanese names won't match English drugs table, but try anyway
        if name_lower in drug_index:
            drug_id = drug_index[name_lower]
            linked += 1

        rows.append({
            "generic_name": r["generic_name"],
            "brand_name": r["brand_name"] or None,
            "strength": r["strength"] or None,
            "registration_number": r["yj_code"] or None,
            "registration_status": "active" if r["status_key"] != "discontinued" else "discontinued",
            "source_name": SOURCE_NAME,
            "source_country": SOURCE_COUNTRY,
            "sponsor": r["manufacturer"] or None,
            "drug_id": drug_id,
        })

    log.info(f"Inserting {len(rows)} catalogue rows ({linked} linked to drugs)...")
    BATCH = 200
    for i in range(0, len(rows), BATCH):
        chunk = rows[i:i + BATCH]
        supabase.table("drug_catalogue").insert(chunk).execute()
        if (i + BATCH) % 2000 < BATCH:
            log.info(f"  {min(i + BATCH, len(rows))}/{len(rows)}")
    log.info(f"Catalogue import complete: {len(rows)} rows, {linked} linked")


def import_shortages(records: list[dict], drug_index: dict[str, str]):
    """Insert shortage records for drugs with limited/suspended supply."""
    shortage_records = [r for r in records if r["severity"] is not None]
    log.info(f"Shortage records to process: {len(shortage_records)}")

    inserted = 0
    skipped = 0
    now = datetime.now(timezone.utc).isoformat()

    for r in shortage_records:
        # Build unique shortage_id
        hash_input = f"JP|{r['yj_code']}|{r['generic_name']}|{r['brand_name']}|{r['status_key']}"
        shortage_id = hashlib.md5(hash_input.encode()).hexdigest()

        # Try to find drug_id
        drug_id = drug_index.get(r["generic_name"].lower().strip())

        # Skip if no drug_id (Japanese names won't match English drugs table)
        # But still insert — the shortage_id dedup will prevent duplicates
        reason_text = r["reason_raw"] if r["reason_raw"] and r["reason_raw"] != "７．ー" else None

        event = {
            "shortage_id": shortage_id,
            "drug_id": drug_id,
            "data_source_id": DATA_SOURCE_ID,
            "country": "Japan",
            "country_code": "JP",
            "status": "active",
            "severity": r["severity"],
            "reason": reason_text,
            "reason_category": r["reason_category"],
            "availability_status": r["availability_status"],
            "management_action": r["resolution_prospect"] if r["resolution_prospect"] and r["resolution_prospect"] != "エ． －" else None,
            "estimated_resolution_date": r["resolution_date"],
            "source_url": INDEX_URL,
            "notes": f"MHLW supply status: {r['availability_status']}. Product: {r['brand_name']}. Manufacturer: {r['manufacturer']}.",
            "raw_data": {
                "generic_name": r["generic_name"],
                "brand_name": r["brand_name"],
                "yj_code": r["yj_code"],
                "manufacturer": r["manufacturer"],
                "shipment_status": r["status_key"],
                "reason_raw": r["reason_raw"],
            },
            "last_verified_at": now,
        }

        try:
            supabase.table("shortage_events").upsert(
                event, on_conflict="shortage_id"
            ).execute()
            inserted += 1
        except Exception as e:
            if "duplicate" in str(e).lower():
                skipped += 1
            else:
                log.warning(f"Error inserting {shortage_id}: {e}")
                skipped += 1

        if (inserted + skipped) % 500 == 0:
            log.info(f"  Processed {inserted + skipped}/{len(shortage_records)} (inserted={inserted}, skipped={skipped})")

    log.info(f"Shortage import complete: {inserted} inserted, {skipped} skipped")


def main():
    excel_url = find_excel_url()
    excel_path = download_excel(excel_url)
    records = parse_excel(excel_path)

    if not records:
        log.error("No records parsed — aborting")
        sys.exit(1)

    # Stats
    by_status = {}
    for r in records:
        by_status[r["status_key"]] = by_status.get(r["status_key"], 0) + 1
    log.info("Supply status breakdown:")
    for k, v in sorted(by_status.items(), key=lambda x: -x[1]):
        log.info(f"  {k:20s}: {v:6d}")

    drug_index = build_drug_index()

    import_catalogue(records, drug_index)
    import_shortages(records, drug_index)

    log.info("All done.")


if __name__ == "__main__":
    main()
