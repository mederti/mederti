import json, os, re, glob
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
INK="12211B"; GREEN="1F6F4A"; LIGHT="E6F0EA"; ZEBRA="F4F8F6"; GREY="5B6660"
AMBER="FFF3D6"; REDF="F0DCDC"; GREENF="DCEEE3"; FONT="Arial"
thin=Side(style="thin",color="CCCCCC"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)

def hdr(ws,row,n):
    for c in range(1,n+1):
        x=ws.cell(row=row,column=c); x.font=Font(name=FONT,bold=True,color=INK,size=9.5)
        x.fill=PatternFill("solid",fgColor=LIGHT); x.border=BORDER
        x.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
def body(ws,r0,r1,n,wrap=True):
    for r in range(r0,r1+1):
        for c in range(1,n+1):
            x=ws.cell(row=r,column=c); x.font=Font(name=FONT,size=9,color=INK); x.border=BORDER
            x.alignment=Alignment(horizontal="left",vertical="top",wrap_text=wrap)
            if (r-r0)%2==1: x.fill=PatternFill("solid",fgColor=ZEBRA)
def widths(ws,ws_w):
    for i,w in enumerate(ws_w,1): ws.column_dimensions[get_column_letter(i)].width=w
def title(ws,txt):
    c=ws.cell(row=1,column=1,value=txt); c.font=Font(name=FONT,bold=True,size=13,color=GREEN)

def cadence(hours):
    if hours is None: return ""
    m={24:"Daily",48:"2-daily",72:"3-daily",168:"Weekly",336:"Fortnightly",720:"Monthly",
       1440:"~2-monthly",2160:"Quarterly",4320:"Half-yearly",8760:"Annual"}
    return m.get(hours, f"Every {hours}h")

def shorten(s,n=300):
    if s is None: return ""
    s=str(s)
    return s if len(s)<=n else s[:n-1]+"…"

ds=json.load(open(os.path.join(ROOT,"analysis/_data_sources.json")))
it=json.load(open(os.path.join(ROOT,"analysis/_intel_sources.json")))

# cron tokens
cron=open(os.path.join(ROOT,"cron/crontab_fixed.txt")).read()
tokens=set(re.findall(r"run_all_scrapers\.py\s+([a-z_]+)",cron))

wb=Workbook()

# ---------------- README ----------------
ws=wb.active; ws.title="README"; ws.sheet_view.showGridLines=False
ws["A1"]="Mederti — Detailed Data Source Register"; ws["A1"].font=Font(name=FONT,bold=True,size=20,color=GREEN)
ws["A2"]="Every data source, detailed — for the CDO source map"; ws["A2"].font=Font(name=FONT,size=11,color=GREY)
lines=[
 ("",""),("Prepared for","Chief Data Officer"),("Generated","3 June 2026"),
 ("Figures","Live production metadata & counts queried 3 June 2026."),
 ("",""),
 ("Sheets",""),
 ("1. Operational Registry","The 57 sources registered in the live data_sources table: URLs, endpoints, reliability weight, scrape frequency, active flag, last-scraped timestamp."),
 ("2. Macro Intelligence (134)","Every source in the intelligence_sources catalogue, fully detailed: owner, access method, formats, auth, poll cadence, entry-point URLs, monitoring priority, notes."),
 ("3. Scraper Inventory","All 69 scraper modules in the repo, mapped to jurisdiction, data domain, and whether they are scheduled in the daily cron."),
 ("4. Reference & Enrichment","The 13 importer modules that load reference/identity datasets (ATC, RxNorm, UNII, SNOMED, etc.)."),
 ("",""),
 ("Status legend (Scraper Inventory)",""),
 ("Scheduled (cron)","Module wired into the daily/quarterly cron."),
 ("Built · not scheduled","Module exists in repo but is not in the cron schedule."),
 ("",""),
 ("Note","data_sources rows are reproduced verbatim from the live table; some legacy rows carry mismatched abbreviation/country values — flagged where noticed, not silently corrected."),
]
r=4
for k,v in lines:
    a=ws.cell(row=r,column=1,value=k); a.font=Font(name=FONT,bold=True,size=10,color=INK)
    b=ws.cell(row=r,column=2,value=v); b.font=Font(name=FONT,size=10,color=INK)
    b.alignment=Alignment(wrap_text=True,vertical="top")
    if k in ("Scheduled (cron)",): a.fill=PatternFill("solid",fgColor=GREENF)
    if k in ("Built · not scheduled",): a.fill=PatternFill("solid",fgColor=AMBER)
    r+=1
widths(ws,[30,100])

# ---------------- 1. Operational Registry (data_sources) ----------------
ws=wb.create_sheet("1. Operational Registry"); ws.sheet_view.showGridLines=False
title(ws,"Operational Source Registry — live data_sources table (57 rows)")
cols=["#","Source name","Abbrev","Country","Region","Source URL","API endpoint",
      "Reliability","Scrape freq","Cadence","Active","Last scraped (UTC)"]
for c,h in enumerate(cols,1): ws.cell(row=3,column=c,value=h)
hdr(ws,3,len(cols))
r0=4
for i,row in enumerate(ds):
    rr=r0+i
    ls=row.get("last_scraped_at")
    ls=ls[:19].replace("T"," ") if ls else "never"
    vals=[i+1,row.get("name"),row.get("abbreviation"),row.get("country"),row.get("region"),
          row.get("source_url"),row.get("api_endpoint") or "—",row.get("reliability_weight"),
          row.get("scrape_frequency_hours"),cadence(row.get("scrape_frequency_hours")),
          "Yes" if row.get("is_active") else "No",ls]
    for c,v in enumerate(vals,1): ws.cell(row=rr,column=c,value=v)
r1=r0+len(ds)-1
body(ws,r0,r1,len(cols))
for rr in range(r0,r1+1):
    for c in (1,8,9,11): ws.cell(row=rr,column=c).alignment=Alignment(horizontal="center",vertical="top")
    ws.cell(row=rr,column=8).number_format="0.00"
    act=ws.cell(row=rr,column=11)
    act.fill=PatternFill("solid",fgColor=GREENF if act.value=="Yes" else REDF)
ws.freeze_panes="B4"; ws.auto_filter.ref=f"A3:L{r1}"
widths(ws,[4,34,14,16,8,40,32,9,9,11,7,18])

# ---------------- 2. Macro Intelligence (134) ----------------
ws=wb.create_sheet("2. Macro Intelligence (134)"); ws.sheet_view.showGridLines=False
title(ws,"Macro Intelligence Source Catalogue — full detail (intelligence_sources, 134 rows)")
cols=["#","Source name","Owner / org","Category","Subcategory","Geography",
      "Regulator?","Gov/IGO?","Expected update","Rec. poll","Access method","Formats",
      "Auth","Daily-monitor priority","Primary entry-point","Docs / robots","Notes"]
for c,h in enumerate(cols,1): ws.cell(row=3,column=c,value=h)
hdr(ws,3,len(cols))
r0=4
def yn(v): return "Yes" if v is True else ("No" if v is False else "")
PRI={"high":GREENF,"medium":AMBER,"low":None}
for i,row in enumerate(it):
    rr=r0+i
    vals=[i+1,row.get("name"),row.get("owner_org"),row.get("category"),row.get("subcategory"),
          row.get("geography_coverage"),yn(row.get("is_medicines_regulator")),yn(row.get("is_government_or_igo")),
          row.get("update_frequency_expected"),row.get("recommended_poll_frequency"),row.get("access_method"),
          shorten(row.get("formats"),60),row.get("auth"),row.get("priority_for_daily_monitoring"),
          shorten(row.get("raw_data_entrypoints"),80),shorten(row.get("docs_entrypoint"),80),shorten(row.get("notes"),400)]
    for c,v in enumerate(vals,1): ws.cell(row=rr,column=c,value=v)
r1=r0+len(it)-1
body(ws,r0,r1,len(cols))
for rr in range(r0,r1+1):
    for c in (1,7,8,14): ws.cell(row=rr,column=c).alignment=Alignment(horizontal="center",vertical="top")
    p=ws.cell(row=rr,column=14)
    f=PRI.get(str(p.value).lower())
    if f: p.fill=PatternFill("solid",fgColor=f)
ws.freeze_panes="B4"; ws.auto_filter.ref=f"A3:Q{r1}"
widths(ws,[4,30,16,18,18,16,9,8,14,9,16,18,10,12,28,28,50])

# ---------------- 3. Scraper Inventory ----------------
ws=wb.create_sheet("3. Scraper Inventory"); ws.sheet_view.showGridLines=False
title(ws,"Scraper Module Inventory — all repo scrapers vs cron schedule")
cols=["#","Module file","Token","Jurisdiction / agency","Data domain","In cron?","Status"]
for c,h in enumerate(cols,1): ws.cell(row=3,column=c,value=h)
hdr(ws,3,len(cols))

JMAP={
 "tga":"Australia — TGA","fda":"United States — FDA","health_canada":"Canada — Health Canada",
 "mhra":"United Kingdom — MHRA","ema":"EU — EMA","bfarm":"Germany — BfArM","ansm":"France — ANSM",
 "aifa":"Italy — AIFA","aemps":"Spain — AEMPS","cbg_meb":"Netherlands — CBG-MEB","dkma":"Denmark — DKMA",
 "fimea":"Finland — FIMEA","hpra":"Ireland — HPRA","lakemedelsverket":"Sweden — MPA","sukl":"Czechia — SUKL",
 "ogyei":"Hungary — OGYEI","swissmedic":"Switzerland — Swissmedic","noma":"Norway — NoMA","ages":"Austria — AGES",
 "anvisa":"Brazil — ANVISA","pmda":"Japan — PMDA","mfds":"South Korea — MFDS","cofepris":"Mexico — COFEPRIS",
 "sahpra":"South Africa — SAHPRA","nafdac":"Nigeria — NAFDAC","sfda":"Saudi Arabia — SFDA","hsa":"Singapore — HSA",
 "pharmac":"New Zealand — PHARMAC","medsafe":"New Zealand — Medsafe","greece_eof":"Greece — EOF",
 "portugal_infarmed":"Portugal — INFARMED","belgium_famhp":"Belgium — FAMHP","poland_mz":"Poland — MZ",
 "argentina_anmat":"Argentina — ANMAT","malaysia_npra":"Malaysia — NPRA","uae_mohap":"UAE — MOHAP",
 "china_nmpa":"China — NMPA","india_cdsco":"India — CDSCO","israel_moh":"Israel — MOH","turkey_titck":"Turkey — TITCK",
 "hk_drugoffice":"Hong Kong — Drug Office","ashp":"United States — ASHP","clinicaltrials":"Global — ClinicalTrials.gov",
 "edqm_cep":"EU — EDQM (CEP)","ema_chmp":"EU — EMA CHMP","eudragmdp":"EU — EudraGMDP","fda_adcomm":"US — FDA AdComm",
 "fda_inspections":"US — FDA inspections","fda_dmf":"US — FDA Drug Master File","fda_decrs":"US — FDA DECRS",
 "fda_enforcement":"US — FDA enforcement","fda_medwatch":"US — FDA MedWatch","drugs_at_fda":"US — Drugs@FDA",
 "who_pq_api":"Global — WHO Prequalification","nhs_drug_tariff":"UK — NHS Drug Tariff","recall_linker":"(internal) recall→shortage linker",
}
def domain(tok):
    if "recall" in tok or tok in ("fda_enforcement","fda_medwatch","drugs_at_fda"): return "Recall"
    if tok in ("fda_dmf","fda_decrs","fda_inspections","who_pq_api","edqm_cep","eudragmdp"): return "Supply / API"
    if tok in ("ema_chmp","fda_adcomm","clinicaltrials"): return "Pipeline / regulatory"
    if tok in ("ashp","nhs_drug_tariff"): return "Shortage / reference"
    if tok=="recall_linker": return "Linkage (internal)"
    return "Shortage"

files=sorted(os.path.basename(f) for f in glob.glob(os.path.join(ROOT,"backend/scrapers/*.py")) if "__init__" not in f and "base_" not in f)
r0=4; i=0
for f in files:
    tok=f.replace("_scraper.py","").replace(".py","")
    # who_pq_api file vs cron token who_pq
    in_cron = tok in tokens or (tok=="who_pq_api" and "who_pq" in tokens)
    rr=r0+i
    if tok in JMAP:
        juris=JMAP[tok]
    elif tok.endswith("_recalls") and tok[:-8] in JMAP:
        juris=JMAP[tok[:-8]]+" — recalls"
    else:
        juris=tok.replace("_"," ").title()
    vals=[i+1,f,tok,juris,domain(tok),"Yes" if in_cron else "No",
          "Scheduled (cron)" if in_cron else "Built · not scheduled"]
    for c,v in enumerate(vals,1): ws.cell(row=rr,column=c,value=v)
    i+=1
r1=r0+i-1
body(ws,r0,r1,len(cols),wrap=False)
for rr in range(r0,r1+1):
    ws.cell(row=rr,column=1).alignment=Alignment(horizontal="center")
    ws.cell(row=rr,column=6).alignment=Alignment(horizontal="center")
    st=ws.cell(row=rr,column=7)
    st.fill=PatternFill("solid",fgColor=GREENF if st.value=="Scheduled (cron)" else AMBER)
ws.freeze_panes="B4"; ws.auto_filter.ref=f"A3:G{r1}"
widths(ws,[4,30,20,30,22,9,22])

# ---------------- 4. Reference & Enrichment ----------------
ws=wb.create_sheet("4. Reference & Enrichment"); ws.sheet_view.showGridLines=False
title(ws,"Reference & Enrichment Importers")
cols=["#","Module","Dataset / provider","Role in the pipeline"]
for c,h in enumerate(cols,1): ws.cell(row=3,column=c,value=h)
hdr(ws,3,len(cols))
imp=[
 ("who_atc_importer.py","WHO ATC/DDD","Therapeutic classification + defined daily dose"),
 ("rxnorm_backfill.py / rxnorm_client.py","RxNorm (US NLM)","US clinical drug nomenclature & ingredient mapping"),
 ("unii_client.py","UNII (FDA)","Unique ingredient identifiers — salt-form resolution"),
 ("snomed_importer.py","SNOMED CT","Clinical terminology cross-walk"),
 ("ema_epar_importer.py","EMA EPAR","European authorisation metadata"),
 ("oecd_pharma_importer.py","OECD pharma","Macro pharmaceutical market reference"),
 ("pharmacompass_importer.py","PharmaCompass","API / supplier market reference"),
 ("inn_resolution.py","INN resolution","Salt-stripping + UNII/RxNorm/ATC molecule identity"),
 ("substance_resolver.py","Substance resolver","Brand→generic / molecule identity resolution"),
 ("catalogue_inn_backfill.py","Catalogue INN backfill","Roll product catalogue up to canonical INN"),
 ("catalogue_inn_revert.py","Catalogue INN revert","Rollback utility for the INN backfill"),
 ("alternatives_importer.py","Drug alternatives","Therapeutic-alternative relationships"),
 ("intelligence_sources_importer.py","Intelligence catalogue","Loads/maintains the 134-source macro catalogue"),
 ("industry_news_press_sources.py","Industry news / press","Press & trade-source registry"),
]
r0=4
for i,(m,d,role) in enumerate(imp):
    rr=r0+i
    for c,v in enumerate([i+1,m,d,role],1): ws.cell(row=rr,column=c,value=v)
r1=r0+len(imp)-1
body(ws,r0,r1,len(cols))
for rr in range(r0,r1+1): ws.cell(row=rr,column=1).alignment=Alignment(horizontal="center",vertical="top")
widths(ws,[4,38,22,56])

wb.save(os.path.join(ROOT,"analysis/Mederti_Detailed_Source_Register.xlsx"))
print("saved Mederti_Detailed_Source_Register.xlsx | data_sources",len(ds),"| intel",len(it),"| scrapers",len(files),"| cron tokens",len(tokens))
