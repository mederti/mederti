"""
Poland Ministry of Health — Threatened Drug Availability List Scraper
─────────────────────────────────────────────────────────────────────
Source:  Ministry of Health Poland — Lista leków zagrożonych brakiem dostępności
URL:     https://www.gov.pl/web/zdrowie/lista-lekow-zagrozonych-brakiem-dostepnosci

The Polish Ministry of Health publishes a list of drugs with threatened
availability ("lista leków zagrożonych brakiem dostępności") on gov.pl.
The page may contain an HTML table or a downloadable PDF/Excel attachment
with shortage data.

Key Polish terms:
    brak dostępności    = lack of availability → status: active
    zagrożony           = threatened           → status: anticipated
    dostępny            = available             → status: resolved
    substancja czynna   = active substance (INN)
    nazwa handlowa      = trade name
    powód               = reason
    data                = date

Data source UUID:  10000000-0000-0000-0000-000000000049
Country:           Poland
Country code:      PL
Confidence:        83/100 (official government source)

Cron:  Every 24 hours (MOH updates infrequently)
"""

from __future__ import annotations

import re
from datetime import date, datetime, timezone
from typing import Any

import httpx
from bs4 import BeautifulSoup

from backend.scrapers.base_scraper import BaseScraper, ScraperError
from backend.utils.reason_mapper import map_reason_category


class PolandMZScraper(BaseScraper):
    SOURCE_ID: str    = "10000000-0000-0000-0000-000000000049"
    SOURCE_NAME: str  = "Ministry of Health Poland — Threatened Drug Availability List"
    BASE_URL: str     = "https://www.gov.pl/web/zdrowie/lista-lekow-zagrozonych-brakiem-dostepnosci"
    COUNTRY: str      = "Poland"
    COUNTRY_CODE: str = "PL"

    RATE_LIMIT_DELAY: float = 2.0
    REQUEST_TIMEOUT: float  = 30.0
    SCRAPER_VERSION: str    = "1.0.0"

    # Polish status keywords → canonical status
    _STATUS_MAP: dict[str, str] = {
        "brak dostępności":   "active",
        "brak dostepnosci":   "active",
        "niedostępny":        "active",
        "niedostepny":        "active",
        "zagrożony":          "anticipated",
        "zagrozony":          "anticipated",
        "zagrożone":          "anticipated",
        "zagrozone":          "anticipated",
        "dostępny":           "resolved",
        "dostepny":           "resolved",
        "przywrócony":        "resolved",
        "przywrocony":        "resolved",
    }

    # Polish reason keywords → canonical reason_category
    _REASON_MAP: dict[str, str] = {
        "produkcja":                   "manufacturing_issue",
        "produkcji":                   "manufacturing_issue",
        "wytwarzanie":                 "manufacturing_issue",
        "jakość":                      "manufacturing_issue",
        "jakosc":                      "manufacturing_issue",
        "surowiec":                    "raw_material",
        "surowce":                     "raw_material",
        "substancja czynna":           "raw_material",
        "substancji czynnej":          "raw_material",
        "popyt":                       "demand_surge",
        "zwiększone zapotrzebowanie":  "demand_surge",
        "zwiekszony popyt":            "demand_surge",
        "łańcuch dostaw":              "supply_chain",
        "lancuch dostaw":              "supply_chain",
        "globalny niedobór":           "supply_chain",
        "globalny niedobor":           "supply_chain",
        "import":                      "distribution",
        "dystrybucja":                 "distribution",
        "wycofanie":                   "discontinuation",
        "wstrzymanie":                 "discontinuation",
        "decyzja handlowa":            "discontinuation",
        "regulacyj":                   "regulatory_action",
        "inspekcja":                   "regulatory_action",
    }

    # -------------------------------------------------------------------------
    # fetch()
    # -------------------------------------------------------------------------

    def fetch(self) -> list[dict]:
        """
        Fetch the Polish MOH threatened availability list page.

        Strategy:
        1. GET the page, parse HTML with BeautifulSoup.
        2. Look for HTML tables with drug shortage data.
        3. Also look for downloadable file links (Excel/PDF attachments).
        """
        self.log.info("Scrape started", extra={
            "source": self.SOURCE_NAME,
            "url": self.BASE_URL,
        })

        try:
            resp = self._get(self.BASE_URL)
            html = resp.text
        except (httpx.TimeoutException, httpx.ConnectTimeout, httpx.ConnectError) as exc:
            self.log.warning(
                "Poland MZ connection/timeout error",
                extra={"error": str(exc), "url": self.BASE_URL},
            )
            return []
        except httpx.HTTPStatusError as exc:
            self.log.warning(
                "Poland MZ HTTP error",
                extra={"status": exc.response.status_code, "url": self.BASE_URL},
            )
            return []

        self.log.info(
            "Poland MZ page fetched",
            extra={"bytes": len(html), "status": resp.status_code},
        )

        soup = BeautifulSoup(html, "html.parser")

        # Try extracting records from HTML tables first
        records = self._parse_html_tables(soup)
        if records:
            self.log.info(
                "Poland MZ HTML table records extracted",
                extra={"records": len(records)},
            )
            return records

        # Try finding downloadable file links (Excel/CSV)
        records = self._parse_download_links(soup)
        if records:
            self.log.info(
                "Poland MZ download link records extracted",
                extra={"records": len(records)},
            )
            return records

        # No data found — log and return empty
        self.log.warning(
            "Poland MZ: no shortage table or download link found on page",
            extra={"url": self.BASE_URL, "html_length": len(html)},
        )
        return []

    def _parse_html_tables(self, soup: BeautifulSoup) -> list[dict]:
        """
        Parse HTML tables from the page for drug shortage data.

        Expected columns may include (Polish):
            Lp., Nazwa, Substancja czynna, Postać, Dawka, Status, Powód, Data
        """
        tables = soup.find_all("table")
        if not tables:
            return []

        records: list[dict] = []
        for table in tables:
            rows = table.find_all("tr")
            if len(rows) < 2:
                continue

            # Extract headers from first row
            header_row = rows[0]
            headers: list[str] = []
            for th in header_row.find_all(["th", "td"]):
                headers.append(th.get_text(strip=True).lower())

            if not headers:
                continue

            # Check if this looks like a drug shortage table
            header_text = " ".join(headers)
            if not any(kw in header_text for kw in (
                "lek", "nazwa", "substancja", "dostępn", "dostepn", "zagroż", "zagroz"
            )):
                continue

            # Parse data rows
            for row in rows[1:]:
                cells = row.find_all(["td", "th"])
                if not cells:
                    continue
                rec: dict[str, str] = {}
                for i, cell in enumerate(cells):
                    key = headers[i] if i < len(headers) else f"col_{i}"
                    rec[key] = cell.get_text(strip=True)
                # Only include rows with some content
                if any(v.strip() for v in rec.values()):
                    records.append(rec)

        return records

    def _parse_download_links(self, soup: BeautifulSoup) -> list[dict]:
        """
        Look for downloadable Excel/CSV attachments on the page and fetch them.
        """
        links = soup.find_all("a", href=True)
        for link in links:
            href = link["href"]
            link_text = link.get_text(strip=True).lower()
            # Look for Excel or CSV file links
            if any(ext in href.lower() for ext in (".xlsx", ".xls", ".csv")):
                return self._fetch_excel_attachment(href)
            if any(kw in link_text for kw in ("pobierz", "download", "lista", "wykaz")):
                if any(ext in href.lower() for ext in (".xlsx", ".xls", ".csv")):
                    return self._fetch_excel_attachment(href)
        return []

    def _fetch_excel_attachment(self, url: str) -> list[dict]:
        """
        Download and parse an Excel attachment linked from the MOH page.
        """
        try:
            import io
            import openpyxl

            # Handle relative URLs
            if url.startswith("/"):
                url = f"https://www.gov.pl{url}"

            self.log.info("Fetching Poland MZ Excel attachment", extra={"url": url})
            resp = self._get(url)

            wb = openpyxl.load_workbook(
                io.BytesIO(resp.content),
                read_only=True,
                data_only=True,
            )
            ws = wb.active

            # Read headers from row 1
            headers: list[str] = []
            for cell in ws[1]:
                headers.append(str(cell.value or "").strip())

            records: list[dict] = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rec = {}
                for h, v in zip(headers, row):
                    if h:
                        rec[h] = v
                if any(v for v in rec.values() if v is not None and str(v).strip()):
                    records.append(rec)

            wb.close()
            self.log.info(
                "Poland MZ Excel parsed",
                extra={"records": len(records), "headers": headers},
            )
            return records

        except Exception as exc:
            self.log.warning(
                "Poland MZ Excel attachment fetch/parse failed",
                extra={"error": str(exc), "url": url},
            )
            return []

    # -------------------------------------------------------------------------
    # normalize()
    # -------------------------------------------------------------------------

    def normalize(self, raw: list[dict]) -> list[dict]:
        """Normalize Poland MZ records into standard shortage event dicts."""
        self.log.info(
            "Normalising Poland MZ records",
            extra={"source": self.SOURCE_NAME, "raw_count": len(raw)},
        )

        normalised: list[dict] = []
        skipped = 0
        today = date.today().isoformat()

        for rec in raw:
            try:
                result = self._normalise_record(rec, today)
                if result is None:
                    skipped += 1
                    continue
                normalised.append(result)
            except Exception as exc:
                skipped += 1
                self.log.warning(
                    "Failed to normalise Poland MZ record",
                    extra={"error": str(exc), "rec": str(rec)[:200]},
                )

        self.log.info(
            "Normalisation done",
            extra={"total": len(raw), "normalised": len(normalised), "skipped": skipped},
        )
        return normalised

    def _normalise_record(self, rec: dict, today: str) -> dict | None:
        """Convert a single Poland MZ record to a normalised shortage event dict."""
        # -- Drug name extraction --
        # Try multiple possible Polish column names
        generic_name = (
            rec.get("substancja czynna")
            or rec.get("substancja")
            or rec.get("nazwa międzynarodowa")
            or rec.get("nazwa miedzynarodowa")
            or rec.get("inn")
            or rec.get("nazwa")
            or rec.get("lek")
            or rec.get("Substancja czynna")
            or rec.get("Nazwa")
            or ""
        )
        if isinstance(generic_name, str):
            generic_name = generic_name.strip()
        else:
            generic_name = str(generic_name).strip()

        if not generic_name:
            return None

        # -- Trade / brand name --
        trade_name = (
            rec.get("nazwa handlowa")
            or rec.get("nazwa produktu")
            or rec.get("nazwa leku")
            or rec.get("Nazwa handlowa")
            or ""
        )
        if isinstance(trade_name, str):
            trade_name = trade_name.strip()
        else:
            trade_name = str(trade_name).strip()

        brand_names = [trade_name] if trade_name and trade_name != generic_name else []

        # -- Shortage reason --
        raw_reason = (
            rec.get("powód")
            or rec.get("powod")
            or rec.get("przyczyna")
            or rec.get("Powód")
            or rec.get("uwagi")
            or ""
        )
        if isinstance(raw_reason, str):
            raw_reason = raw_reason.strip()
        else:
            raw_reason = str(raw_reason).strip()

        reason_category = self._map_reason(raw_reason)

        # -- Start date --
        raw_start = (
            rec.get("data")
            or rec.get("data zgłoszenia")
            or rec.get("data zgloszenia")
            or rec.get("data od")
            or rec.get("Data")
        )
        start_date = self._parse_date(raw_start) or today

        # -- Status --
        raw_status = (
            rec.get("status")
            or rec.get("Status")
            or rec.get("dostępność")
            or rec.get("dostepnosc")
            or ""
        )
        status = self._map_status(str(raw_status))

        # -- Dosage / form info for notes --
        dosage = str(rec.get("dawka") or rec.get("Dawka") or "").strip()
        form = str(rec.get("postać") or rec.get("postac") or rec.get("Postać") or "").strip()

        # -- Build notes --
        notes_parts: list[str] = []
        if form:
            notes_parts.append(f"Form: {form}")
        if dosage:
            notes_parts.append(f"Dosage: {dosage}")
        if raw_reason:
            notes_parts.append(f"Reason (PL): {raw_reason}")
        notes = "; ".join(notes_parts) or None

        return {
            "generic_name":              generic_name.title(),
            "brand_names":               brand_names,
            "status":                    status,
            "severity":                  "medium",
            "reason":                    raw_reason or None,
            "reason_category":           reason_category,
            "start_date":                start_date,
            "source_url":                self.BASE_URL,
            "notes":                     notes,
            "source_confidence_score":   83,
            "raw_record":                rec,
        }

    # -------------------------------------------------------------------------
    # Private helpers
    # -------------------------------------------------------------------------

    def _map_status(self, raw: str) -> str:
        """Map Polish status string to canonical status."""
        if not raw:
            return "anticipated"  # Default: items on the list are threatened
        lower = raw.strip().lower()
        for key, status in self._STATUS_MAP.items():
            if key in lower:
                return status
        # Default for items on the threatened availability list
        return "anticipated"

    def _map_reason(self, raw: str) -> str:
        """Map Polish reason string to canonical reason_category."""
        if not raw:
            return "unknown"
        lower = raw.strip().lower()
        for key, cat in self._REASON_MAP.items():
            if key in lower:
                return cat
        # Fallback to centralized mapper
        return map_reason_category(raw)

    @staticmethod
    def _parse_date(raw: Any) -> str | None:
        """Parse various date formats to ISO-8601 date string."""
        if raw is None:
            return None
        if isinstance(raw, datetime):
            return raw.date().isoformat()
        if isinstance(raw, date):
            return raw.isoformat()
        raw_str = str(raw).strip()
        if not raw_str or raw_str in ("-", "N/A", "null", "None", ""):
            return None

        # Try Polish date format: DD.MM.YYYY or DD-MM-YYYY
        for pattern, fmt in [
            (r"^\d{2}\.\d{2}\.\d{4}$", "%d.%m.%Y"),
            (r"^\d{2}-\d{2}-\d{4}$",   "%d-%m-%Y"),
            (r"^\d{4}-\d{2}-\d{2}$",   "%Y-%m-%d"),
            (r"^\d{2}/\d{2}/\d{4}$",   "%d/%m/%Y"),
        ]:
            if re.match(pattern, raw_str):
                try:
                    return datetime.strptime(raw_str, fmt).date().isoformat()
                except ValueError:
                    pass

        # Fallback to dateutil
        try:
            from dateutil import parser as dtparser
            dt = dtparser.parse(raw_str, dayfirst=True)
            return dt.date().isoformat()
        except (ValueError, ImportError):
            pass
        return None


# -------------------------------------------------------------------------
# Standalone entrypoint
# -------------------------------------------------------------------------

if __name__ == "__main__":
    import json
    import os
    import sys

    from dotenv import load_dotenv

    load_dotenv()

    dry_run = os.environ.get("MEDERTI_DRY_RUN", "0").strip() == "1"

    if dry_run:
        from unittest.mock import MagicMock

        print("=" * 60)
        print("DRY RUN MODE  (MEDERTI_DRY_RUN=1)")
        print("Fetches live Poland MZ data but makes NO database writes.")
        print("Set MEDERTI_DRY_RUN=0 to run against Supabase.")
        print("=" * 60)

        scraper = PolandMZScraper(db_client=MagicMock())

        print("\n-- Fetching from Poland MZ ...")
        raw = scraper.fetch()
        print(f"-- Raw records received : {len(raw)}")

        print("-- Normalising records ...")
        events = scraper.normalize(raw)
        print(f"-- Normalised events    : {len(events)}")

        if events:
            print("\n-- Sample event (first record, raw_record omitted):")
            sample = {k: v for k, v in events[0].items() if k != "raw_record"}
            print(json.dumps(sample, indent=2, default=str))

            from collections import Counter

            status_counts   = Counter(e["status"] for e in events)
            severity_counts = Counter(e.get("severity") for e in events)
            reason_counts   = Counter(e.get("reason_category") for e in events)

            print("\n-- Status breakdown:")
            for k, v in sorted(status_counts.items()):
                print(f"   {k:25s} {v}")
            print("\n-- Severity breakdown:")
            for k, v in sorted(severity_counts.items()):
                print(f"   {str(k):12s} {v}")
            print("\n-- Reason category breakdown:")
            for k, v in sorted(reason_counts.items()):
                print(f"   {str(k):30s} {v}")

        print("\n-- Dry run complete. No writes made to Supabase.")
        sys.exit(0)

    # -- Live run --
    print("=" * 60)
    print("LIVE RUN - writing to Supabase")
    print("=" * 60)

    scraper = PolandMZScraper()
    summary = scraper.run()

    print("\n-- Scrape summary:")
    print(json.dumps(summary, indent=2, default=str))

    sys.exit(0 if summary["status"] in ("success", "duplicate") else 1)
