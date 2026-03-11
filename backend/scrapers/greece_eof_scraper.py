"""
Greece EOF National Organisation for Medicines — Drug Shortages Scraper
───────────────────────────────────────────────────────────────────────
Source:  EOF — Ethnikos Organismos Farmakon (National Organisation for Medicines)
URL:     https://www.eof.gr/

The Greek National Organisation for Medicines (EOF) publishes drug shortage
notifications and alerts on its website. EOF is the Hellenic equivalent of
the EMA at national level. Shortage notices may appear as announcements,
downloadable PDFs, or HTML tables under the /web/guest/notifications section.

Key Greek terms:
    έλλειψη / elleipsi      = shortage           → status: active
    ανάκληση / anaklisi      = recall             → status: active
    αναστολή / anastoli      = suspension         → status: active
    διαθέσιμο / diathesimo   = available          → status: resolved
    αναμένεται / anamenete   = expected/pending   → status: anticipated
    φάρμακο / farmako        = drug/medicine
    δραστική ουσία           = active substance (INN)
    εμπορική ονομασία        = trade name

Data source UUID:  10000000-0000-0000-0000-000000000050
Country:           Greece
Country code:      GR
Confidence:        82/100 (official national regulator)

Cron:  Every 24 hours (EOF updates infrequently)
"""

from __future__ import annotations

import re
from datetime import date, datetime, timezone
from typing import Any

import httpx
from bs4 import BeautifulSoup

from backend.scrapers.base_scraper import BaseScraper, ScraperError
from backend.utils.reason_mapper import map_reason_category


class GreeceEOFScraper(BaseScraper):
    SOURCE_ID: str    = "10000000-0000-0000-0000-000000000050"
    SOURCE_NAME: str  = "National Organisation for Medicines — Drug Shortages"
    BASE_URL: str     = "https://www.eof.gr/"
    COUNTRY: str      = "Greece"
    COUNTRY_CODE: str = "GR"

    RATE_LIMIT_DELAY: float = 2.0
    REQUEST_TIMEOUT: float  = 30.0
    SCRAPER_VERSION: str    = "1.0.0"

    # Common EOF sub-pages for shortage/notification data
    _SHORTAGE_PATHS: list[str] = [
        "web/guest/notifications",
        "web/guest/shortages",
        "web/guest/ellipseis",
        "web/guest/announcements",
    ]

    # Greek status keywords → canonical status
    _STATUS_MAP: dict[str, str] = {
        "έλλειψη":      "active",
        "ελλειψη":      "active",
        "elleipsi":     "active",
        "ανάκληση":     "active",
        "ανακληση":     "active",
        "αναστολή":     "active",
        "αναστολη":     "active",
        "μη διαθέσιμο": "active",
        "μη διαθεσιμο": "active",
        "διαθέσιμο":    "resolved",
        "διαθεσιμο":    "resolved",
        "αναμένεται":   "anticipated",
        "αναμενεται":   "anticipated",
        "προσωρινή":    "anticipated",
        "προσωρινη":    "anticipated",
    }

    # Greek reason keywords → canonical reason_category
    _REASON_MAP: dict[str, str] = {
        "παραγωγή":                "manufacturing_issue",
        "παραγωγη":                "manufacturing_issue",
        "κατασκευή":               "manufacturing_issue",
        "κατασκευη":               "manufacturing_issue",
        "ποιότητα":                "manufacturing_issue",
        "ποιοτητα":                "manufacturing_issue",
        "πρώτη ύλη":               "raw_material",
        "πρωτη υλη":               "raw_material",
        "δραστική ουσία":          "raw_material",
        "δραστικη ουσια":          "raw_material",
        "ζήτηση":                  "demand_surge",
        "ζητηση":                  "demand_surge",
        "αυξημένη ζήτηση":        "demand_surge",
        "αυξημενη ζητηση":        "demand_surge",
        "εφοδιαστική αλυσίδα":    "supply_chain",
        "εφοδιαστικη αλυσιδα":    "supply_chain",
        "παγκόσμια έλλειψη":      "supply_chain",
        "παγκοσμια ελλειψη":      "supply_chain",
        "εισαγωγή":               "distribution",
        "εισαγωγη":               "distribution",
        "διανομή":                 "distribution",
        "διανομη":                 "distribution",
        "απόσυρση":                "discontinuation",
        "αποσυρση":                "discontinuation",
        "διακοπή κυκλοφορίας":    "discontinuation",
        "διακοπη κυκλοφοριας":    "discontinuation",
        "ρυθμιστικ":               "regulatory_action",
        "κανονιστικ":              "regulatory_action",
    }

    # -------------------------------------------------------------------------
    # fetch()
    # -------------------------------------------------------------------------

    def fetch(self) -> list[dict]:
        """
        Fetch EOF shortage/notification pages.

        Strategy:
        1. GET the base EOF URL and known sub-pages for shortage data.
        2. Parse HTML with BeautifulSoup looking for tables or structured lists.
        3. Aggregate all shortage-related records found.
        """
        self.log.info("Scrape started", extra={
            "source": self.SOURCE_NAME,
            "url": self.BASE_URL,
        })

        all_records: list[dict] = []

        # Try main page first
        records = self._fetch_page(self.BASE_URL)
        all_records.extend(records)

        # Try known sub-pages for shortage data
        for path in self._SHORTAGE_PATHS:
            url = f"{self.BASE_URL.rstrip('/')}/{path}"
            try:
                page_records = self._fetch_page(url)
                all_records.extend(page_records)
            except Exception as exc:
                self.log.debug(
                    f"EOF sub-page not accessible: {path}",
                    extra={"error": str(exc)},
                )

        self.log.info(
            "EOF fetch complete",
            extra={"total_records": len(all_records)},
        )
        return all_records

    def _fetch_page(self, url: str) -> list[dict]:
        """Fetch a single EOF page and extract shortage records."""
        try:
            resp = self._get(url)
            html = resp.text
        except (httpx.TimeoutException, httpx.ConnectTimeout, httpx.ConnectError) as exc:
            self.log.warning(
                "EOF connection/timeout error",
                extra={"error": str(exc), "url": url},
            )
            return []
        except httpx.HTTPStatusError as exc:
            self.log.warning(
                "EOF HTTP error",
                extra={"status": exc.response.status_code, "url": url},
            )
            return []

        self.log.debug(
            "EOF page fetched",
            extra={"bytes": len(html), "status": resp.status_code, "url": url},
        )

        soup = BeautifulSoup(html, "html.parser")

        # Try HTML tables
        records = self._parse_html_tables(soup, url)
        if records:
            return records

        # Try structured announcement lists
        records = self._parse_announcement_list(soup, url)
        if records:
            return records

        # Try downloadable attachments (Excel/CSV)
        records = self._parse_download_links(soup, url)
        if records:
            return records

        return []

    def _parse_html_tables(self, soup: BeautifulSoup, source_url: str) -> list[dict]:
        """
        Parse HTML tables for drug shortage data.

        Expected columns may include (Greek):
            Α/Α, Ονομασία, Δραστική Ουσία, Μορφή, Δοσολογία, Κατάσταση, Αιτία
        """
        tables = soup.find_all("table")
        if not tables:
            return []

        records: list[dict] = []
        for table in tables:
            rows = table.find_all("tr")
            if len(rows) < 2:
                continue

            # Extract headers from first row
            header_row = rows[0]
            headers: list[str] = []
            for th in header_row.find_all(["th", "td"]):
                headers.append(th.get_text(strip=True).lower())

            if not headers:
                continue

            # Check if this looks like a drug shortage table (Greek or English)
            header_text = " ".join(headers)
            if not any(kw in header_text for kw in (
                "φάρμακο", "φαρμακο", "ονομασία", "ονομασια",
                "δραστική", "δραστικη", "έλλειψ", "ελλειψ",
                "drug", "medicine", "shortage", "substance",
            )):
                continue

            # Parse data rows
            for row in rows[1:]:
                cells = row.find_all(["td", "th"])
                if not cells:
                    continue
                rec: dict[str, str] = {"_source_url": source_url}
                for i, cell in enumerate(cells):
                    key = headers[i] if i < len(headers) else f"col_{i}"
                    rec[key] = cell.get_text(strip=True)
                if any(v.strip() for k, v in rec.items() if k != "_source_url"):
                    records.append(rec)

        return records

    def _parse_announcement_list(self, soup: BeautifulSoup, source_url: str) -> list[dict]:
        """
        Parse structured announcement/notification lists for shortage data.
        EOF may publish shortage notifications as news items or announcements.
        """
        records: list[dict] = []

        # Look for announcement containers with shortage-related content
        shortage_keywords = [
            "έλλειψη", "ελλειψη", "shortage", "ανάκληση", "ανακληση",
            "αναστολή", "αναστολη", "φάρμακο", "φαρμακο",
        ]

        # Search in article/news containers
        for container in soup.find_all(["article", "div", "li"], class_=True):
            text = container.get_text(strip=True).lower()
            if any(kw in text for kw in shortage_keywords):
                # Extract title/heading
                title_el = container.find(["h2", "h3", "h4", "a", "strong"])
                if title_el:
                    title = title_el.get_text(strip=True)
                    # Extract date if present
                    date_el = container.find(["time", "span"], class_=lambda c: c and "date" in str(c).lower())
                    date_text = date_el.get_text(strip=True) if date_el else ""

                    records.append({
                        "title":       title,
                        "date":        date_text,
                        "full_text":   container.get_text(strip=True)[:500],
                        "_source_url": source_url,
                    })

        return records

    def _parse_download_links(self, soup: BeautifulSoup, source_url: str) -> list[dict]:
        """
        Look for downloadable Excel/CSV attachments on the page.
        """
        links = soup.find_all("a", href=True)
        for link in links:
            href = link["href"]
            link_text = link.get_text(strip=True).lower()
            # Look for Excel or CSV file links
            if any(ext in href.lower() for ext in (".xlsx", ".xls", ".csv")):
                return self._fetch_excel_attachment(href)
            # Greek download keywords
            if any(kw in link_text for kw in ("λήψη", "ληψη", "κατέβασμα", "κατεβασμα", "download")):
                if any(ext in href.lower() for ext in (".xlsx", ".xls", ".csv")):
                    return self._fetch_excel_attachment(href)
        return []

    def _fetch_excel_attachment(self, url: str) -> list[dict]:
        """Download and parse an Excel attachment linked from the EOF page."""
        try:
            import io
            import openpyxl

            # Handle relative URLs
            if url.startswith("/"):
                url = f"https://www.eof.gr{url}"

            self.log.info("Fetching EOF Excel attachment", extra={"url": url})
            resp = self._get(url)

            wb = openpyxl.load_workbook(
                io.BytesIO(resp.content),
                read_only=True,
                data_only=True,
            )
            ws = wb.active

            # Read headers from row 1
            headers: list[str] = []
            for cell in ws[1]:
                headers.append(str(cell.value or "").strip())

            records: list[dict] = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rec = {}
                for h, v in zip(headers, row):
                    if h:
                        rec[h] = v
                if any(v for v in rec.values() if v is not None and str(v).strip()):
                    records.append(rec)

            wb.close()
            self.log.info(
                "EOF Excel parsed",
                extra={"records": len(records), "headers": headers},
            )
            return records

        except Exception as exc:
            self.log.warning(
                "EOF Excel attachment fetch/parse failed",
                extra={"error": str(exc), "url": url},
            )
            return []

    # -------------------------------------------------------------------------
    # normalize()
    # -------------------------------------------------------------------------

    def normalize(self, raw: list[dict]) -> list[dict]:
        """Normalize EOF records into standard shortage event dicts."""
        self.log.info(
            "Normalising EOF records",
            extra={"source": self.SOURCE_NAME, "raw_count": len(raw)},
        )

        normalised: list[dict] = []
        skipped = 0
        today = date.today().isoformat()

        for rec in raw:
            try:
                result = self._normalise_record(rec, today)
                if result is None:
                    skipped += 1
                    continue
                normalised.append(result)
            except Exception as exc:
                skipped += 1
                self.log.warning(
                    "Failed to normalise EOF record",
                    extra={"error": str(exc), "rec": str(rec)[:200]},
                )

        self.log.info(
            "Normalisation done",
            extra={"total": len(raw), "normalised": len(normalised), "skipped": skipped},
        )
        return normalised

    def _normalise_record(self, rec: dict, today: str) -> dict | None:
        """Convert a single EOF record to a normalised shortage event dict."""
        source_url = rec.pop("_source_url", self.BASE_URL)

        # -- Drug name extraction --
        # Try multiple possible Greek column names
        generic_name = (
            rec.get("δραστική ουσία")
            or rec.get("δραστικη ουσια")
            or rec.get("δραστική")
            or rec.get("δραστικη")
            or rec.get("ονομασία")
            or rec.get("ονομασια")
            or rec.get("active substance")
            or rec.get("substance")
            or rec.get("inn")
            or rec.get("title")
            or rec.get("name")
            or ""
        )
        if isinstance(generic_name, str):
            generic_name = generic_name.strip()
        else:
            generic_name = str(generic_name).strip()

        if not generic_name:
            return None

        # If name is from announcement title, try to extract drug name
        if len(generic_name) > 100:
            # Likely a full announcement — try to extract drug reference
            extracted = self._extract_drug_from_text(generic_name)
            if extracted:
                generic_name = extracted
            else:
                return None  # Cannot extract meaningful drug name

        # -- Trade / brand name --
        trade_name = (
            rec.get("εμπορική ονομασία")
            or rec.get("εμπορικη ονομασια")
            or rec.get("trade name")
            or rec.get("brand")
            or ""
        )
        if isinstance(trade_name, str):
            trade_name = trade_name.strip()
        else:
            trade_name = str(trade_name).strip()

        brand_names = [trade_name] if trade_name and trade_name != generic_name else []

        # -- Shortage reason --
        raw_reason = (
            rec.get("αιτία")
            or rec.get("αιτια")
            or rec.get("λόγος")
            or rec.get("λογος")
            or rec.get("reason")
            or rec.get("cause")
            or ""
        )
        if isinstance(raw_reason, str):
            raw_reason = raw_reason.strip()
        else:
            raw_reason = str(raw_reason).strip()

        reason_category = self._map_reason(raw_reason)

        # -- Start date --
        raw_start = (
            rec.get("ημερομηνία")
            or rec.get("ημερομηνια")
            or rec.get("date")
            or rec.get("ημ/νία")
            or rec.get("ημ/νια")
        )
        start_date = self._parse_date(raw_start) or today

        # -- Status --
        raw_status = (
            rec.get("κατάσταση")
            or rec.get("κατασταση")
            or rec.get("status")
            or rec.get("full_text")
            or ""
        )
        status = self._map_status(str(raw_status))

        # -- Dosage / form info for notes --
        dosage = str(rec.get("δοσολογία") or rec.get("δοσολογια") or rec.get("dosage") or "").strip()
        form = str(rec.get("μορφή") or rec.get("μορφη") or rec.get("form") or "").strip()

        # -- Build notes --
        notes_parts: list[str] = []
        if form:
            notes_parts.append(f"Form: {form}")
        if dosage:
            notes_parts.append(f"Dosage: {dosage}")
        if raw_reason:
            notes_parts.append(f"Reason (GR): {raw_reason}")
        notes = "; ".join(notes_parts) or None

        return {
            "generic_name":              generic_name.title(),
            "brand_names":               brand_names,
            "status":                    status,
            "severity":                  "medium",
            "reason":                    raw_reason or None,
            "reason_category":           reason_category,
            "start_date":                start_date,
            "source_url":                source_url,
            "notes":                     notes,
            "source_confidence_score":   82,
            "raw_record":                rec,
        }

    # -------------------------------------------------------------------------
    # Private helpers
    # -------------------------------------------------------------------------

    def _map_status(self, raw: str) -> str:
        """Map Greek status string to canonical status."""
        if not raw:
            return "active"
        lower = raw.strip().lower()
        for key, status in self._STATUS_MAP.items():
            if key in lower:
                return status
        return "active"

    def _map_reason(self, raw: str) -> str:
        """Map Greek reason string to canonical reason_category."""
        if not raw:
            return "unknown"
        lower = raw.strip().lower()
        for key, cat in self._REASON_MAP.items():
            if key in lower:
                return cat
        # Fallback to centralized mapper
        return map_reason_category(raw)

    @staticmethod
    def _extract_drug_from_text(text: str) -> str | None:
        """
        Attempt to extract a drug/substance name from a longer announcement text.
        Looks for patterns like Latin INN names (capitalized single/double words)
        that are commonly used even in Greek announcements.
        """
        if not text:
            return None

        # INN names are typically Latin-script words within Greek text
        # Look for capitalized Latin words of sufficient length
        latin_words = re.findall(r'\b[A-Z][a-zA-Z]{3,}\b', text)
        if latin_words:
            # Return the first substantial Latin-script word as likely INN
            return latin_words[0]

        return None

    @staticmethod
    def _parse_date(raw: Any) -> str | None:
        """Parse various date formats to ISO-8601 date string."""
        if raw is None:
            return None
        if isinstance(raw, datetime):
            return raw.date().isoformat()
        if isinstance(raw, date):
            return raw.isoformat()
        raw_str = str(raw).strip()
        if not raw_str or raw_str in ("-", "N/A", "null", "None", ""):
            return None

        # Try Greek/European date formats: DD/MM/YYYY, DD.MM.YYYY, DD-MM-YYYY
        for pattern, fmt in [
            (r"^\d{2}/\d{2}/\d{4}$", "%d/%m/%Y"),
            (r"^\d{2}\.\d{2}\.\d{4}$", "%d.%m.%Y"),
            (r"^\d{2}-\d{2}-\d{4}$",   "%d-%m-%Y"),
            (r"^\d{4}-\d{2}-\d{2}$",   "%Y-%m-%d"),
        ]:
            if re.match(pattern, raw_str):
                try:
                    return datetime.strptime(raw_str, fmt).date().isoformat()
                except ValueError:
                    pass

        # Fallback to dateutil
        try:
            from dateutil import parser as dtparser
            dt = dtparser.parse(raw_str, dayfirst=True)
            return dt.date().isoformat()
        except (ValueError, ImportError):
            pass
        return None


# -------------------------------------------------------------------------
# Standalone entrypoint
# -------------------------------------------------------------------------

if __name__ == "__main__":
    import json
    import os
    import sys

    from dotenv import load_dotenv

    load_dotenv()

    dry_run = os.environ.get("MEDERTI_DRY_RUN", "0").strip() == "1"

    if dry_run:
        from unittest.mock import MagicMock

        print("=" * 60)
        print("DRY RUN MODE  (MEDERTI_DRY_RUN=1)")
        print("Fetches live EOF Greece data but makes NO database writes.")
        print("Set MEDERTI_DRY_RUN=0 to run against Supabase.")
        print("=" * 60)

        scraper = GreeceEOFScraper(db_client=MagicMock())

        print("\n-- Fetching from EOF Greece ...")
        raw = scraper.fetch()
        print(f"-- Raw records received : {len(raw)}")

        print("-- Normalising records ...")
        events = scraper.normalize(raw)
        print(f"-- Normalised events    : {len(events)}")

        if events:
            print("\n-- Sample event (first record, raw_record omitted):")
            sample = {k: v for k, v in events[0].items() if k != "raw_record"}
            print(json.dumps(sample, indent=2, default=str))

            from collections import Counter

            status_counts   = Counter(e["status"] for e in events)
            severity_counts = Counter(e.get("severity") for e in events)
            reason_counts   = Counter(e.get("reason_category") for e in events)

            print("\n-- Status breakdown:")
            for k, v in sorted(status_counts.items()):
                print(f"   {k:25s} {v}")
            print("\n-- Severity breakdown:")
            for k, v in sorted(severity_counts.items()):
                print(f"   {str(k):12s} {v}")
            print("\n-- Reason category breakdown:")
            for k, v in sorted(reason_counts.items()):
                print(f"   {str(k):30s} {v}")

        print("\n-- Dry run complete. No writes made to Supabase.")
        sys.exit(0)

    # -- Live run --
    print("=" * 60)
    print("LIVE RUN - writing to Supabase")
    print("=" * 60)

    scraper = GreeceEOFScraper()
    summary = scraper.run()

    print("\n-- Scrape summary:")
    print(json.dumps(summary, indent=2, default=str))

    sys.exit(0 if summary["status"] in ("success", "duplicate") else 1)
