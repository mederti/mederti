"""
Japan MHLW Drug Supply Status Scraper
──────────────────────────────────────
Source:  Ministry of Health, Labour and Welfare (MHLW)
URL:     https://www.mhlw.go.jp/stf/seisakunitsuite/bunya/kenkou_iryou/iryou/kouhatu-iyaku/04_00003.html

Data access
───────────
MHLW publishes the full pharmaceutical supply status as an Excel file,
updated roughly weekly. The file lists all ~16,500 drug products with
their supply status:

    ①通常出荷                    Normal supply
    ②限定出荷（自社の事情）       Limited shipment (company reasons)
    ③限定出荷（他社品の影響）     Limited shipment (other companies)
    ④限定出荷（その他）           Limited shipment (other)
    ⑤供給停止                    Supply suspended

Only non-normal statuses (②–⑤) are imported as shortage events.

Key columns:
    Col 3  (③成分名)              Generic name (INN)
    Col 6  (⑥品名)               Product/brand name
    Col 7  (⑦製造販売業者名)      Manufacturer
    Col 4  (④規格単位)            Strength/dosage form
    Col 5  (⑤YJコード)            YJ code (Japanese drug code)
    Col 12 (⑫出荷対応)            Supply status
    Col 13 (⑬更新日)              Status update date
    Col 14 (⑭理由)                Reason for shortage

Data source UUID:  10000000-0000-0000-0000-000000000037  (MHLW, JP)
Country:           Japan
Country code:      JP
"""

from __future__ import annotations

import io
import re
from datetime import datetime, timezone
from typing import Any

import httpx

from backend.scrapers.base_scraper import BaseScraper


class PmdaScraper(BaseScraper):
    """Scraper for MHLW Japan drug supply status Excel file."""

    SOURCE_ID:    str = "10000000-0000-0000-0000-000000000037"
    SOURCE_NAME:  str = "Ministry of Health, Labour and Welfare — Drug Supply Status"
    BASE_URL:     str = "https://www.mhlw.go.jp/stf/seisakunitsuite/bunya/kenkou_iryou/iryou/kouhatu-iyaku/04_00003.html"
    COUNTRY:      str = "Japan"
    COUNTRY_CODE: str = "JP"

    RATE_LIMIT_DELAY: float = 2.0
    REQUEST_TIMEOUT:  float = 60.0
    SCRAPER_VERSION:  str   = "2.0.0"

    _HEADERS: dict = {
        "User-Agent":      "Mozilla/5.0 (compatible; Mederti-Scraper/2.0)",
        "Accept-Language": "ja,en;q=0.8",
    }

    # Supply status mapping (Japanese → Mederti status + severity)
    _STATUS_MAP: dict[str, tuple[str, str]] = {
        "①通常出荷":             ("resolved", "low"),        # Normal supply
        "②限定出荷（自社の事情）": ("active", "medium"),      # Limited - company reasons
        "③限定出荷（他社品の影響）": ("active", "medium"),    # Limited - other companies
        "④限定出荷（その他）":     ("active", "medium"),      # Limited - other
        "⑤供給停止":             ("active", "critical"),      # Supply suspended
    }

    # Reason mapping
    _REASON_CODES: dict[str, str] = {
        "１": "manufacturing_issue",       # 製造上の問題
        "２": "raw_material",              # 原材料の調達困難
        "３": "demand_surge",              # 需要の増加
        "４": "regulatory_action",         # 行政措置
        "５": "supply_chain",              # その他の供給問題
        "６": "discontinuation",           # 薬価削除予定
        "７": "unknown",                   # その他
    }

    # ─────────────────────────────────────────────────────────────────────────
    # fetch()
    # ─────────────────────────────────────────────────────────────────────────

    def fetch(self) -> list[dict]:
        """
        1. GET the MHLW index page to find the current Excel URL.
        2. Download the Excel file.
        3. Parse rows, filtering to only non-normal supply statuses.
        """
        from bs4 import BeautifulSoup

        self.log.info("Fetching MHLW index page", extra={"url": self.BASE_URL})

        # Step 1: Find the Excel link
        with httpx.Client(
            headers=self._HEADERS,
            timeout=self.REQUEST_TIMEOUT,
            follow_redirects=True,
        ) as client:
            resp = client.get(self.BASE_URL)
            resp.raise_for_status()

        soup = BeautifulSoup(resp.text, "html.parser")
        xlsx_url = None
        for a in soup.find_all("a", href=True):
            if ".xlsx" in a["href"].lower():
                href = a["href"]
                xlsx_url = href if href.startswith("http") else f"https://www.mhlw.go.jp{href}"
                break

        if not xlsx_url:
            self.log.error("MHLW: no Excel link found on index page")
            return []

        self.log.info("Downloading MHLW Excel", extra={"url": xlsx_url})

        # Step 2: Download the Excel
        with httpx.Client(
            headers=self._HEADERS,
            timeout=self.REQUEST_TIMEOUT,
            follow_redirects=True,
        ) as client:
            resp = client.get(xlsx_url)
            resp.raise_for_status()

        # Step 3: Parse the Excel
        return self._parse_xlsx(resp.content)

    def _parse_xlsx(self, content: bytes) -> list[dict]:
        """Parse the MHLW supply status Excel, returning only shortage records."""
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active or wb[wb.sheetnames[0]]

        records: list[dict] = []
        skipped_normal = 0

        for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
            if not row or len(row) < 14:
                continue

            # Column indices (0-based from the Excel):
            # 0=薬剤区分, 1=薬効分類, 2=成分名, 3=規格単位, 4=YJコード
            # 5=品名, 6=製造販売業者名, 7=製品区分, 8=基礎的医薬品
            # 9=供給確保医薬品, 10=薬価収載年月日, 11=出荷対応, 12=更新日, 13=理由
            supply_status = str(row[11] or "").strip()

            if not supply_status:
                continue

            # Skip normal supply products
            if supply_status == "①通常出荷":
                skipped_normal += 1
                continue

            generic_name = str(row[2] or "").strip()
            if not generic_name or len(generic_name) < 2:
                continue

            brand_name = str(row[5] or "").strip()
            manufacturer = str(row[6] or "").strip()
            strength = str(row[3] or "").strip()
            yj_code = str(row[4] or "").strip()
            drug_category = str(row[0] or "").strip()
            therapeutic_class = str(row[1] or "").strip()
            product_type = str(row[7] or "").strip()
            update_date = row[12]
            reason_raw = str(row[13] or "").strip()

            # Convert datetime to string for JSON serialization
            update_date_str = ""
            if isinstance(update_date, datetime):
                update_date_str = update_date.date().isoformat()
            elif update_date:
                update_date_str = str(update_date).strip()

            records.append({
                "generic_name": generic_name,
                "brand_name": brand_name,
                "manufacturer": manufacturer,
                "strength": strength,
                "yj_code": yj_code,
                "drug_category": drug_category,
                "therapeutic_class": therapeutic_class,
                "product_type": product_type,
                "supply_status": supply_status,
                "update_date": update_date_str,
                "reason_raw": reason_raw,
            })

        self.log.info(
            "MHLW Excel parsed",
            extra={
                "total_rows": i + 1 if 'i' in dir() else 0,
                "shortage_records": len(records),
                "normal_skipped": skipped_normal,
            },
        )
        wb.close()
        return records

    # ─────────────────────────────────────────────────────────────────────────
    # normalize()
    # ─────────────────────────────────────────────────────────────────────────

    def normalize(self, raw: list[dict]) -> list[dict]:
        """Normalize MHLW records into standard shortage event dicts."""
        self.log.info("Normalising MHLW records", extra={"raw_count": len(raw)})

        today = datetime.now(timezone.utc).date().isoformat()
        normalised: list[dict] = []
        skipped = 0

        for rec in raw:
            try:
                result = self._normalise_record(rec, today)
                if result is None:
                    skipped += 1
                    continue
                normalised.append(result)
            except Exception as exc:
                skipped += 1
                self.log.warning(
                    "MHLW: normalise error",
                    extra={"error": str(exc), "rec": str(rec)[:200]},
                )

        self.log.info(
            "MHLW normalisation done",
            extra={"normalised": len(normalised), "skipped": skipped},
        )
        return normalised

    def _normalise_record(self, rec: dict, today: str) -> dict | None:
        """Convert a single MHLW record to standard format."""
        generic_name = rec.get("generic_name", "").strip()
        if not generic_name:
            return None

        brand_name = rec.get("brand_name", "").strip()
        brand_names = [brand_name] if brand_name and brand_name != generic_name else []

        # Status + severity from supply status
        supply_status = rec.get("supply_status", "")
        status, severity = self._STATUS_MAP.get(supply_status, ("active", "medium"))

        # Skip resolved (normal supply) — should already be filtered
        if status == "resolved":
            return None

        # Reason mapping
        reason_raw = rec.get("reason_raw", "")
        reason_category = "unknown"
        for code, cat in self._REASON_CODES.items():
            if reason_raw.startswith(code) or f"．{code}" in reason_raw:
                reason_category = cat
                break

        # Reason text
        reason_text = supply_status
        if "供給停止" in supply_status:
            reason_text = "Supply suspended"
        elif "限定出荷" in supply_status:
            if "自社" in supply_status:
                reason_text = "Limited shipment (company reasons)"
            elif "他社" in supply_status:
                reason_text = "Limited shipment (impact from other companies)"
            else:
                reason_text = "Limited shipment (other)"

        # Parse update date
        update_date = rec.get("update_date", "")
        start_date = today
        if update_date:
            ud = str(update_date).strip()
            # Try ISO format (YYYY-MM-DD)
            iso_match = re.match(r"^(\d{4})-(\d{2})-(\d{2})", ud)
            if iso_match:
                start_date = iso_match.group(0)
            # Try datetime string with time component
            elif "T" in ud or " " in ud:
                dt_match = re.match(r"^(\d{4})-(\d{2})-(\d{2})", ud)
                if dt_match:
                    start_date = dt_match.group(0)

        # Notes
        notes_parts: list[str] = []
        manufacturer = rec.get("manufacturer", "").strip()
        if manufacturer:
            notes_parts.append(f"Manufacturer: {manufacturer}")
        strength = rec.get("strength", "").strip()
        if strength:
            notes_parts.append(f"Strength: {strength}")
        yj_code = rec.get("yj_code", "").strip()
        if yj_code:
            notes_parts.append(f"YJ: {yj_code}")
        therapeutic = rec.get("therapeutic_class", "").strip()
        if therapeutic:
            notes_parts.append(f"Class: {therapeutic}")
        if reason_raw:
            notes_parts.append(f"Reason code: {reason_raw}")
        notes = "; ".join(notes_parts) or None

        return {
            "generic_name":              generic_name,
            "brand_names":               brand_names,
            "status":                    status,
            "severity":                  severity,
            "reason":                    reason_text,
            "reason_category":           reason_category,
            "start_date":                start_date,
            "source_url":                self.BASE_URL,
            "notes":                     notes,
            "source_confidence_score":   90,
            "raw_record":                rec,
        }


if __name__ == "__main__":
    import json, os, sys
    from dotenv import load_dotenv
    load_dotenv()
    dry_run = os.environ.get("MEDERTI_DRY_RUN", "0").strip() == "1"
    if dry_run:
        from unittest.mock import MagicMock
        from collections import Counter
        print("=" * 60); print("DRY RUN — MHLW Japan Supply Status"); print("=" * 60)
        scraper = PmdaScraper(db_client=MagicMock())
        raw = scraper.fetch()
        print(f"  Raw shortage records: {len(raw)}")
        events = scraper.normalize(raw)
        print(f"  Normalised events  : {len(events)}")
        if events:
            sample = {k: v for k, v in events[0].items() if k != "raw_record"}
            print(f"\n-- Sample event:")
            print(json.dumps(sample, indent=2, ensure_ascii=False, default=str))

            status_counts = Counter(e["status"] for e in events)
            severity_counts = Counter(e.get("severity") for e in events)
            reason_counts = Counter(e.get("reason_category") for e in events)

            print("\n-- Status breakdown:")
            for k, v in sorted(status_counts.items()):
                print(f"   {k:25s} {v}")
            print("\n-- Severity breakdown:")
            for k, v in sorted(severity_counts.items()):
                print(f"   {str(k):12s} {v}")
            print("\n-- Reason category breakdown:")
            for k, v in sorted(reason_counts.items()):
                print(f"   {str(k):30s} {v}")

        print("\n-- Dry run complete.")
        sys.exit(0)

    scraper = PmdaScraper()
    summary = scraper.run()
    print(json.dumps(summary, indent=2, default=str))
    sys.exit(0 if summary["status"] in ("success", "duplicate") else 1)
