"""
EMA EPAR Importer (Path A — 3/3)
─────────────────────────────────
Pulls the European Medicines Agency's public bulk download of centrally-
authorized medicines (the "EPAR" table) and patches `drugs.ema_product_number`
with each drug's EMEA/H/C/xxxxxx procedure number.

Why this matters
────────────────
The EMEA/H/C/ number is the canonical reference EU regulators, procurement
teams, and EPAR documents use to identify a centrally-authorized medicine.
Mederti already shows ATC + INN; adding the EMA product number means a
hospital pharmacist looking at our atorvastatin page can paste the number
straight into the EMA EPAR search to grab the SmPC.

Coverage
────────
Only centrally-authorized products carry EMEA/H/C/ numbers. Nationally-
authorized drugs (the majority of the EU market) won't get one — that's
expected, not a bug. Each Mederti drug row will either pick up an EMA
number or stay null.

Matching
────────
We lowercase + strip both the EMA "International non-proprietary name (INN) /
common name" column and `drugs.generic_name`, then look up direct matches.
First-match wins per INN — when multiple branded centrally-authorized
products share an INN, we take the first EMEA number seen. The schema
column is TEXT for now; if we need to keep the full list later we can
migrate to TEXT[] without breaking callers.

We NEVER clobber an existing `ema_product_number` value — if a drug row
already has one set (e.g. by a manual review), we skip.

Source
──────
EMA publishes the EPAR Excel as a public bulk download. URL has moved
historically — we try the current canonical path first and surface a
clear error if it 404s, so the next agent can update the constant.

Usage
─────
    python3 -m backend.importers.ema_epar_importer                # full run
    python3 -m backend.importers.ema_epar_importer --limit 20     # smoke
    MEDERTI_DRY_RUN=1 python3 -m backend.importers.ema_epar_importer

Cadence
───────
Quarterly is plenty — centrally-authorized products don't churn fast.
Idempotent (we only fill nulls), so re-running is cheap.

Schema dependency
─────────────────
Requires migration 035 (`drugs.ema_product_number` column). Will fail
loudly if the column doesn't exist.
"""
from __future__ import annotations

import argparse
import io
import os
import re
import sys
from collections import defaultdict
from typing import Any, Iterable

import httpx
from openpyxl import load_workbook

# Sibling imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from backend.utils.db import get_supabase_client  # noqa: E402

DRY_RUN = os.environ.get("MEDERTI_DRY_RUN", "0") == "1"

# EMA's bulk "Medicines" Excel — covers every centrally-authorized product
# with INN, ATC, EPAR procedure number, MAH and authorisation status. EMA
# renamed and re-pathed this file in 2024 (used to be
# "Medicines_output_european_public_assessment_reports.xlsx" under
# /sites/default/files/). If THIS 404s, check
# https://www.ema.europa.eu/en/medicines/download-medicine-data for the
# current link — the index page lives at a stable URL even when files move.
EMA_XLSX_URL = (
    "https://www.ema.europa.eu/en/documents/report/"
    "medicines-output-medicines-report_en.xlsx"
)
EMA_XLSX_URL_FALLBACK = (
    "https://www.ema.europa.eu/sites/default/files/"
    "Medicines_output_european_public_assessment_reports.xlsx"
)

USER_AGENT = "Mederti-EMA-Importer/1.0 (https://mederti.com)"
DOWNLOAD_TIMEOUT_S = 60.0
PAGE_SIZE = 1000

# EMEA/H/C/ numbers are six digits, optionally with /xxxx suffixes for
# variations. We only want the base procedure number.
EMEA_NUMBER_RE = re.compile(r"\bEMEA/H/C/\d{4,7}\b", re.IGNORECASE)

# Header normalisation — EMA's column names have shifted over the years
# ("Medicine name" vs "Name of medicine", "Active substance" sometimes
# splits from INN). We accept any of the known aliases.
INN_COLUMN_ALIASES = {
    "international non-proprietary name (inn) / common name",
    "international non-proprietary name (inn) or common name",
    "international non-proprietary name",
    "active substance",
    "common name",
    "inn",
}
PRODUCT_NUMBER_ALIASES = {
    "product number",
    "ema product number",
    "procedure number",
}
MEDICINE_NAME_ALIASES = {
    "medicine name",
    "name of medicine",
    "name",
}


# ──────────────────────────────────────────────────────────────────────────────
# Download
# ──────────────────────────────────────────────────────────────────────────────

def download_xlsx() -> bytes:
    """Fetch the EMA EPAR Excel. Tries primary URL, then fallback path."""
    headers = {"User-Agent": USER_AGENT, "Accept": "*/*"}
    for url in (EMA_XLSX_URL, EMA_XLSX_URL_FALLBACK):
        try:
            print(f"[EMA EPAR] Downloading {url}…", flush=True)
            r = httpx.get(url, headers=headers, timeout=DOWNLOAD_TIMEOUT_S, follow_redirects=True)
            if r.status_code == 200 and r.content:
                print(f"[EMA EPAR] {len(r.content):,} bytes", flush=True)
                return r.content
            print(f"  ! {url} → HTTP {r.status_code}", flush=True)
        except httpx.HTTPError as e:
            print(f"  ! {url} → {e}", flush=True)
    raise RuntimeError(
        "Could not download the EMA EPAR Excel from either known URL. "
        "Check https://www.ema.europa.eu/en/medicines/download-medicine-data "
        "for the current link and update EMA_XLSX_URL in this module."
    )


# ──────────────────────────────────────────────────────────────────────────────
# Parse
# ──────────────────────────────────────────────────────────────────────────────

def _normalise_header(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def _resolve_headers(header_row: Iterable[Any]) -> dict[str, int]:
    """
    Find the column indices we care about (INN, product number, medicine name).
    Raises if a required column is missing — better to fail loud than write
    garbage into the DB.
    """
    found: dict[str, int] = {}
    for idx, raw in enumerate(header_row):
        norm = _normalise_header(str(raw) if raw is not None else "")
        if "inn_col" not in found and norm in INN_COLUMN_ALIASES:
            found["inn_col"] = idx
        if "product_number_col" not in found and norm in PRODUCT_NUMBER_ALIASES:
            found["product_number_col"] = idx
        if "medicine_name_col" not in found and norm in MEDICINE_NAME_ALIASES:
            found["medicine_name_col"] = idx

    missing = [k for k in ("inn_col", "product_number_col") if k not in found]
    if missing:
        raise RuntimeError(
            f"EMA EPAR Excel missing expected columns: {missing}. "
            f"Saw headers: {[str(r) for r in header_row if r is not None][:25]}"
        )
    return found


def _extract_emea_number(raw: Any) -> str | None:
    """Pull out a clean EMEA/H/C/xxxxxx number from a cell value."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    m = EMEA_NUMBER_RE.search(s)
    if not m:
        return None
    # Normalise casing — EMA uses uppercase canonically.
    return m.group(0).upper()


def parse_xlsx(data: bytes) -> dict[str, str]:
    """
    Walk the workbook and return {inn_lowered: ema_number} for the first
    EMEA number seen per INN. Multiple branded products under the same INN
    collapse to a single first-match — acceptable for the schema column we
    have today.
    """
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)

    # The EMA workbook usually has multiple sheets (one per data tab); the
    # EPAR rows are typically on the first sheet, but its name varies
    # ("epar" / "Sheet1" / "Medicines"). We scan sheets until one has the
    # headers we need.
    by_inn: dict[str, str] = {}
    rows_scanned = 0
    rows_with_number = 0
    rows_with_inn = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Find a header row in the first ~15 rows. EMA sometimes prefixes
        # the data with metadata rows ("Last updated:", blank rows).
        header_idx: int | None = None
        header_row: list[Any] = []
        for r_idx, row in enumerate(ws.iter_rows(values_only=True, max_row=15)):
            norms = {_normalise_header(str(c) if c is not None else "") for c in row}
            if norms & INN_COLUMN_ALIASES and norms & PRODUCT_NUMBER_ALIASES:
                header_idx = r_idx
                header_row = list(row)
                break

        if header_idx is None:
            continue  # try next sheet

        cols = _resolve_headers(header_row)
        inn_col = cols["inn_col"]
        num_col = cols["product_number_col"]
        name_col = cols.get("medicine_name_col")

        print(
            f"[EMA EPAR] Parsing sheet {sheet_name!r}: "
            f"INN col={inn_col} number col={num_col} "
            f"medicine col={name_col}", flush=True
        )

        # Re-iterate from the row after the header.
        for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if r_idx <= header_idx:
                continue
            rows_scanned += 1

            number = _extract_emea_number(row[num_col] if num_col < len(row) else None)
            if number:
                rows_with_number += 1

            raw_inn = row[inn_col] if inn_col < len(row) else None
            if raw_inn is None:
                continue
            # INN cells often list multiple substances comma- or
            # semicolon-separated (combination products). We index by every
            # listed INN so a Mederti drug for any of them can pick up the
            # number — better than dropping combinations entirely.
            inn_text = str(raw_inn).strip().lower()
            if not inn_text:
                continue
            rows_with_inn += 1

            if not number:
                continue

            for part in re.split(r"[,;/]", inn_text):
                key = part.strip()
                if not key:
                    continue
                # First-match wins — don't overwrite an INN that's already
                # been mapped to an earlier product number.
                by_inn.setdefault(key, number)

        break  # we found and parsed the right sheet; stop

    print(
        f"[EMA EPAR] rows_scanned={rows_scanned} "
        f"rows_with_inn={rows_with_inn} rows_with_number={rows_with_number} "
        f"unique_inns_mapped={len(by_inn)}", flush=True
    )
    return by_inn


# ──────────────────────────────────────────────────────────────────────────────
# Match + patch
# ──────────────────────────────────────────────────────────────────────────────

def fetch_drugs_needing_ema_number(supabase: Any, limit: int | None = None) -> list[dict[str, Any]]:
    """
    Return drugs with a non-empty generic_name and no ema_product_number set.
    We page through in batches of PAGE_SIZE to avoid 1000-row caps.
    """
    print("[EMA EPAR] Loading drugs needing EMA numbers…", flush=True)
    drugs: list[dict[str, Any]] = []
    offset = 0
    while True:
        page = (
            supabase.table("drugs")
            .select("id, generic_name, ema_product_number")
            .is_("ema_product_number", "null")
            .not_.is_("generic_name", "null")
            .range(offset, offset + PAGE_SIZE - 1)
            .execute()
            .data or []
        )
        drugs.extend(page)
        if len(page) < PAGE_SIZE:
            break
        offset += PAGE_SIZE

    print(f"[EMA EPAR] {len(drugs)} drugs eligible for patching", flush=True)
    if limit:
        drugs = drugs[:limit]
        print(f"[EMA EPAR] --limit {limit} → considering {len(drugs)}", flush=True)
    return drugs


def patch_drug(supabase: Any, drug_id: str, ema_number: str) -> bool:
    if DRY_RUN:
        return True
    try:
        supabase.table("drugs").update({"ema_product_number": ema_number}).eq("id", drug_id).execute()
        return True
    except Exception as e:
        print(f"  ! patch failed for {drug_id}: {e}", flush=True)
        return False


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=None,
                    help="Only process N candidate drugs (smoke testing)")
    ap.add_argument("--skip-download", action="store_true",
                    help="Use a local copy of the EMA Excel (set EMA_EPAR_LOCAL_PATH)")
    args = ap.parse_args()

    # Fetch the Excel.
    if args.skip_download:
        path = os.environ.get("EMA_EPAR_LOCAL_PATH")
        if not path or not os.path.exists(path):
            print("EMA_EPAR_LOCAL_PATH must point to a downloaded EMA EPAR .xlsx", flush=True)
            return 2
        with open(path, "rb") as f:
            data = f.read()
        print(f"[EMA EPAR] Loaded {len(data):,} bytes from {path}", flush=True)
    else:
        data = download_xlsx()

    inn_to_number = parse_xlsx(data)
    if not inn_to_number:
        print("[EMA EPAR] No INN→EMEA mappings extracted — aborting before any DB writes.", flush=True)
        return 1

    supabase = get_supabase_client()
    candidates = fetch_drugs_needing_ema_number(supabase, limit=args.limit)
    if not candidates:
        print("[EMA EPAR] Nothing to patch ✓", flush=True)
        return 0

    matched = 0
    patched = 0
    by_outcome: dict[str, int] = defaultdict(int)
    for i, drug in enumerate(candidates, start=1):
        generic = (drug.get("generic_name") or "").strip().lower()
        if not generic:
            by_outcome["no_generic"] += 1
            continue

        # Direct match first; fall back to comma-splitting (some Mederti
        # rows carry combination products as "drug A / drug B").
        number = inn_to_number.get(generic)
        if not number:
            for part in re.split(r"[,;/]", generic):
                p = part.strip()
                if p and p in inn_to_number:
                    number = inn_to_number[p]
                    break

        if not number:
            by_outcome["no_ema_match"] += 1
            continue

        matched += 1
        if patch_drug(supabase, drug["id"], number):
            patched += 1
            by_outcome["patched"] += 1
        else:
            by_outcome["patch_error"] += 1

        if i % 100 == 0:
            print(f"  …{i}/{len(candidates)}  matched={matched}  patched={patched}", flush=True)

    print(
        f"[EMA EPAR] Done ✓  candidates={len(candidates)}  matched={matched}  "
        f"patched={patched}  outcomes={dict(by_outcome)}",
        flush=True,
    )
    if DRY_RUN:
        print("[EMA EPAR] DRY RUN — no rows written. Re-run without MEDERTI_DRY_RUN to persist.", flush=True)
    return 0


if __name__ == "__main__":
    sys.exit(main())
